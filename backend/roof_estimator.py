"""
Roof Estimator - Drawing-based quantity takeoff and cost estimation.

Takes manual measurements from scaled architectural drawings and calculates
material quantities based on specification requirements (Div 07).

Replaces satellite/footprint-based estimation with spec-driven calculations.

Usage:
    python roof_estimator.py
    python roof_estimator.py --json output.json
"""

import math
import json
import sys
import logging
from dataclasses import dataclass, field

logger = logging.getLogger(__name__)

try:
    import openpyxl
except ImportError:  # Optional dependency for Excel takeoff overrides
    openpyxl = None

from backend.database import (
    PRICING,
    EPDM_SPECIFIC_MATERIALS,
    TPO_SPECIFIC_MATERIALS,
    COMMON_ROOF_MATERIALS,
    ROOF_SYSTEM_CONFIGS,
)


# ---------------------------------------------------------------------------
# Coverage rates - how much area/length one purchase unit covers
# Used by AI-driven detail calculations to convert area -> unit count
# ---------------------------------------------------------------------------

COVERAGE_RATES = {
    "Primer":                          {"sqft_per_unit": 250, "unit": "pail"},
    "Base_Membrane":                   {"sqft_per_unit": 100, "unit": "roll"},
    "Cap_Membrane":                    {"sqft_per_unit": 86,  "unit": "roll"},
    "SBS_Membrane":                    {"sqft_per_unit": 100, "unit": "roll"},
    "EPDM_Membrane":                   {"sqft_per_unit": 1000, "unit": "roll (10'x100')"},
    "TPO_Membrane":                    {"sqft_per_unit": 1000, "unit": "roll (10'x100')"},
    "PVC_Membrane":                    {"sqft_per_unit": 100, "unit": "roll"},
    "Vapour_Barrier_Membrane":         {"sqft_per_unit": 200, "unit": "roll"},
    "EPDM_Accessory":                  {"sqft_per_unit": 50,  "unit": "roll"},
    "TPO_Accessory":                   {"sqft_per_unit": 50,  "unit": "piece"},
    "Polyisocyanurate_ISO_Insulation": {"sqft_per_unit": 16,  "unit": "sheet (4'x4')"},
    "XPS_Insulation":                  {"sqft_per_unit": 16,  "unit": "sheet (2'x8')"},
    "Fiberboard_Insulation":           {"sqft_per_unit": 8,   "unit": "sheet (2'x4')"},
    "Batt_Insulation":                 {"sqft_per_unit": 40,  "unit": "bundle"},
    "DensDeck_Coverboard":             {"sqft_per_unit": 32,  "unit": "sheet (4'x8')"},
    "Gypsum_Fiber_Coverboard":         {"sqft_per_unit": 32,  "unit": "sheet (4'x8')"},
    "Drainage_Board":                  {"sqft_per_unit": 300, "unit": "roll (6'x50')"},
    "Fleece_Reinforcement_Fabric":     {"sqft_per_unit": 300, "unit": "roll"},
    "Flashing_General":                {"lf_per_unit": 10,    "unit": "10ft piece"},
    "Coated_Metal_Sheet":              {"sqft_per_unit": 40,  "unit": "sheet (4'x10')"},
    "Metal_Panel":                     {"lf_per_unit": 1,     "unit": "lin ft"},
    "Standing_Seam_Metal":             {"lf_per_unit": 1,     "unit": "lin ft"},
    "Drip_Edge":                       {"lf_per_unit": 10,    "unit": "10ft piece"},
    "Wood_Blocking_Lumber":            {"lf_per_unit": 8,     "unit": "8ft piece"},
    "Plywood_Sheathing":              {"sqft_per_unit": 32,  "lf_per_unit": 8,  "unit": "4'x8' sheet"},
    "Gravel_Ballast":                  {"sqft_per_unit": 1,   "unit": "sqft"},
    "Filter_Fabric":                   {"sqft_per_unit": 1,   "unit": "sqft"},
    "Mastic":                          {"sqft_per_unit": 500, "unit": "pail"},
    "Adhesive":                        {"sqft_per_unit": 200, "unit": "pail"},
    "Adhesive_Elastocol":              {"sqft_per_unit": 333, "unit": "pail (19L)"},
    "Sealant_General":                 {"lf_per_unit": 20,    "unit": "tube"},
    "Sealant_Mulco_Supra":             {"lf_per_unit": 15,    "unit": "tube"},
    "Coating_Paint":                   {"sqft_per_unit": 200, "unit": "pail"},
    "Tape":                            {"lf_per_unit": 150,   "unit": "roll"},
    "Walkway_Pads":                    {"sqft_per_unit": 12,  "unit": "pad"},
    "Roof_Drain":                      {"per_each": 1, "unit": "EA"},
    "Scupper":                         {"per_each": 1, "unit": "EA"},
    "Gooseneck_Vent":                  {"per_each": 1, "unit": "EA"},
    "Pipe_Boot_Seal":                  {"per_each": 1, "unit": "EA"},
    "Plumbing_Vent":                   {"per_each": 1, "unit": "EA"},
    "Vent_Cap":                        {"per_each": 1, "unit": "EA"},
    "Roof_Hatch":                      {"per_each": 1, "unit": "EA"},
    "Roof_Anchor":                     {"per_each": 1, "unit": "EA"},
    "Gutter_Downpipe":                 {"per_each": 1, "unit": "EA"},
    "Clips":                           {"per_each": 1, "unit": "piece"},
    "Fasteners":                       {"sqft_per_unit": 100, "lf_per_unit": 400, "unit": "box (1M)"},
    "Insulation_Plates":               {"sqft_per_unit": 100, "lf_per_unit": 400, "unit": "box (1M)"},
    "Nails_Staples":                   {"sqft_per_unit": 200, "lf_per_unit": 600, "unit": "box"},
    "Screws":                          {"sqft_per_unit": 100, "lf_per_unit": 400, "unit": "box"},
    "HVAC_Curb_Detail":                {"per_each": 1, "unit": "EA"},
    "Equipment_Torch":                 {"sqft_per_unit": 100, "unit": "roll"},
    # --- EPDM/TPO specific keys used by AI detail path ---
    "EPDM_Membrane_60mil":             {"sqft_per_unit": 1000, "unit": "roll (10'x100')"},
    "EPDM_Membrane_45mil":             {"sqft_per_unit": 1000, "unit": "roll (10'x100')"},
    "EPDM_Filter_Fabric":              {"sqft_per_unit": 300,  "unit": "roll"},
    "EPDM_Drainage_Mat":               {"sqft_per_unit": 300,  "unit": "roll (6'x50')"},
    "EPDM_Seam_Tape":                  {"sqft_per_unit": 1000, "unit": "roll (100 lf)"},
    "EPDM_Bonding_Adhesive":           {"sqft_per_unit": 300,  "unit": "pail (5 gal)"},
    "EPDM_Primer_HP250":               {"sqft_per_unit": 50,   "unit": "gallon"},
    "EPDM_Cav_Grip":                   {"sqft_per_unit": 500,  "unit": "cylinder"},
    "EPDM_Lap_Sealant":                {"lf_per_unit": 22,     "unit": "tube"},
    "EPDM_PS_Corner":                  {"per_each": 1,         "unit": "piece"},
    "EPDM_Curb_Flash":                 {"lf_per_unit": 50,     "unit": "roll"},
    "EPDM_RUSS_6":                     {"lf_per_unit": 100,    "unit": "roll"},
    "EPDM_Pipe_Flashing":              {"per_each": 1,         "unit": "piece"},
    "Vapour_Barrier_Sopravapor":        {"sqft_per_unit": 500,  "unit": "roll"},
    "ISO_2_5_inch":                     {"sqft_per_unit": 16,   "unit": "sheet (4'x4')"},
    "Densdeck_Half_Inch":               {"sqft_per_unit": 32,   "unit": "sheet (4'x8')"},
    "Tapered_ISO":                      {"sqft_per_unit": 1,    "unit": "sqft"},
    "Soprasmart_ISO_HD":                {"sqft_per_unit": 32,   "unit": "sheet (4'x8')"},
    "Duotack_Adhesive":                 {"sqft_per_unit": 500,  "unit": "case"},
    "Elastocol_Stick":                  {"sqft_per_unit": 333,  "unit": "pail (19L)"},
    "TPO_Bonding_Adhesive_SureWeld":    {"sqft_per_unit": 300,  "unit": "pail (5 gal)"},
    "TPO_Primer":                       {"sqft_per_unit": 100,  "unit": "gallon"},
    "TPO_Flashing_24in":                {"lf_per_unit": 50,     "unit": "roll"},
    "TPO_Flashing_12in":                {"lf_per_unit": 50,     "unit": "roll"},
    "TPO_Rhinobond_Plate":              {"sqft_per_unit": 4000, "unit": "pallet"},
    "TPO_Screws":                       {"sqft_per_unit": 4000, "unit": "box"},
    "TPO_Corner":                       {"per_each": 1,         "unit": "piece"},
    "TPO_Tuck_Tape":                    {"lf_per_unit": 150,    "unit": "roll"},
    "TPO_Lap_Sealant":                  {"lf_per_unit": 22,     "unit": "tube"},
    "TPO_Pipe_Boot":                    {"per_each": 1,         "unit": "piece"},
    # --- Common materials ---
    "Vapour_Barrier_TieIn":             {"per_each": 1,         "unit": "allowance"},
    "Fire_Prevention_Board":            {"sqft_per_unit": 20,   "unit": "sheet (20 sqft)"},
    "Screws_Plates_Combo":              {"sqft_per_unit": 1000, "unit": "box (1M)"},
    "Flashing_Bond_Mastic":             {"lf_per_unit": 50,     "unit": "tube"},
    "Flashing_Bond_Mastic_Garland":     {"lf_per_unit": 50,     "unit": "pail"},
    "Asphalt_EasyMelt":                 {"sqft_per_unit": 200,  "unit": "pail"},
    "Catalyst":                         {"per_each": 1,         "unit": "can"},
    "Sopralap_Cover_Strip":             {"lf_per_unit": 75,     "unit": "roll"},
    "Roof_Tape_IKO":                    {"lf_per_unit": 75,     "unit": "roll"},
    "Tuff_Stuff_MS":                    {"lf_per_unit": 15,     "unit": "tube"},
    "Gar_Mesh":                         {"sqft_per_unit": 300,  "unit": "roll"},
    "Garla_Flex":                       {"sqft_per_unit": 200,  "unit": "pail"},
    "PMMA_Primer":                      {"sqft_per_unit": 200,  "unit": "pail (10kg)"},
    "Gum_Box":                          {"per_each": 1,         "unit": "EA"},
    # --- Wood/lumber ---
    "Cant_Strip_4x4":                   {"lf_per_unit": 8,      "unit": "8ft piece"},
    "Lumber_2x4":                       {"lf_per_unit": 8,      "unit": "8ft piece"},
    "Lumber_2x6":                       {"lf_per_unit": 8,      "unit": "8ft piece"},
    "Lumber_2x10":                      {"lf_per_unit": 8,      "unit": "8ft piece"},
    "Plywood_Three_Quarter":            {"sqft_per_unit": 32,   "unit": "4'x8' sheet"},
    "Plywood_Half":                     {"sqft_per_unit": 32,   "unit": "4'x8' sheet"},
    "Metal_Flashing_Galvanized":        {"lf_per_unit": 1,      "unit": "LF"},
    "Metal_Flashing_Prepainted":        {"lf_per_unit": 1,      "unit": "LF"},
    "Metal_Cladding_Panel":             {"sqft_per_unit": 1,    "unit": "sqft"},
    "EPS_Insulation_EPDM":             {"sqft_per_unit": 16,   "unit": "sheet (4'x4')"},
}

# Map AI detail_type -> (measurement_type, RoofMeasurements attribute)
DETAIL_TYPE_MAP = {
    "field_assembly":          ("sqft",      "total_roof_area_sqft"),
    "parapet":                 ("linear_ft", "parapet_length_lf"),
    "curtain_wall":            ("linear_ft", "parapet_length_lf"),
    "drain":                   ("each",      "roof_drain_count"),
    "mechanical_curb":         ("each",      "mechanical_unit_count"),
    "sleeper_curb":            ("each",      "sleeper_curb_count"),
    "penetration_gas":         ("each",      "gas_penetration_count"),
    "penetration_electrical":  ("each",      "electrical_penetration_count"),
    "penetration_plumbing":    ("each",      "plumbing_vent_count"),
    "vent_hood":               ("each",      "vent_hood_count"),
    "scupper":                 ("each",      "scupper_count"),
    "expansion_joint":         ("linear_ft", "perimeter_lf"),
    "pipe_support":            ("each",      "plumbing_vent_count"),
    "opening_cover":           ("each",      "mechanical_unit_count"),
}

# Typical curb perimeters (LF) for fallback mode when AI doesn't provide dimensions
CURB_TYPICAL_PERIMETER_LF = {
    "sleeper_curb": 7,       # ~3' x 0.5' typical → ~7 LF perimeter
    "mechanical_curb": 52,   # ~18' x 8' typical → ~52 LF perimeter
}


def _material_scope(pricing_key: str) -> str:
    """Classify a material as 'area', 'linear', or 'discrete' based on COVERAGE_RATES."""
    cov = COVERAGE_RATES.get(pricing_key, {})
    if "per_each" in cov:
        return "discrete"
    if "lf_per_unit" in cov and "sqft_per_unit" not in cov:
        return "linear"
    return "area"


# ---------------------------------------------------------------------------
# New Takeoff Data Structures (Excel: Takeoff Sheet parity)
# ---------------------------------------------------------------------------

PERIMETER_TYPES = {
    "parapet_no_facing": "Parapet w/o Facing",
    "parapet_w_facing": "Parapet w/ Facing",
    "interior_wall": "Interior Wall",
    "cant": "Cant",
    "divider_w_facing": "Divider w/ Facing",
}

# Vent labour hours lookup (Excel: Takeoff H43-H48)
# Keys: base hours + adjustment per difficulty variant
VENT_LABOUR_HOURS = {
    "pipe_boot":  {"base": 1.0, "Normal": 0.0, "Hard": 1.0},
    "b_vent":     {"base": 3.0, "No_Curb": 0.0, "Curb": 2.0},
    "hood_vent":  {"base": 4.0, "Normal": -1.0, "Hard": 1.0},
    "plumb_vent": {"base": 1.5, "Normal": 0.0, "Hard": 0.5},
    "gum_box":    {"base": 3.0, "Normal": 0.0, "Hard": 1.0},
    "scupper":    {"base": 2.0, "Easy": -0.5, "Normal": 0.0, "Hard": 1.0},
    "radon_pipe": {"base": 1.5, "Normal": 0.0, "Hard": 0.5},
    "drain":      {"base": 2.0, "Drop_Drain": -1.0, "Normal": 0.0, "Mech_Attachment": 1.0},
}

# Metal flashing pricing keys by type
METAL_FLASHING_TYPES = {
    "galvanized":  "Metal_Flashing_Galvanized",
    "prepainted":  "Metal_Flashing_Prepainted",
    "cladding":    "Metal_Cladding_Panel",
}

# Wood product pricing keys (Excel: FRS R115-R120)
WOOD_PRODUCT_KEYS = {
    "cant_4x4":   "Cant_Strip_4x4",
    "lumber_2x4": "Lumber_2x4",
    "lumber_2x6": "Lumber_2x6",
    "lumber_2x10": "Lumber_2x10",
    "plywood_3_4": "Plywood_Three_Quarter",
    "plywood_1_2": "Plywood_Half",
}


@dataclass
class ProjectSettings:
    """Project-level modifiers from the Excel Project sheet.
    These affect perimeter/cladding install hour rates."""
    base_flashing_rate: float = 7.5      # LF/hr (Project!R38)
    floor_count: int = 1                  # Project!D5
    hot_work: bool = False                # Project!D18
    tear_off: bool = False                # Project!D19
    interior_access_only: bool = False    # Project!D20
    winter_conditions: bool = False       # Project!D21
    scaffold_factor: float = 0.85
    hot_work_factor: float = 0.90
    tear_off_factor: float = 0.90
    interior_access_factor: float = 0.85
    winter_factor: float = 0.85

    @property
    def effective_rate(self) -> float:
        """Compute the adjusted flashing install rate after all modifiers."""
        rate = self.base_flashing_rate
        if self.floor_count > 3:
            rate *= self.scaffold_factor
        if self.hot_work:
            rate *= self.hot_work_factor
        if self.tear_off:
            rate *= self.tear_off_factor
        if self.interior_access_only:
            rate *= self.interior_access_factor
        if self.winter_conditions:
            rate *= self.winter_factor
        return rate


@dataclass
class RoofSection:
    """Individual flat roof section (Excel: Takeoff R5-R14).
    Up to 6 sections, each with count x length x width."""
    name: str = ""
    count: int = 1
    length_ft: float = 0.0
    width_ft: float = 0.0

    @property
    def area_sqft(self) -> float:
        return self.count * self.length_ft * self.width_ft


@dataclass
class CurbDetail:
    """Dimensioned curb (Excel: Takeoff R31-R37).
    Types: RTU, Roof_Hatch, Vent_Curb, Skylight."""
    curb_type: str = "RTU"
    count: int = 0
    length_in: float = 48.0
    width_in: float = 48.0
    height_in: float = 18.0

    @property
    def perimeter_lf_each(self) -> float:
        return 2 * (self.length_in + self.width_in) / 12.0

    @property
    def total_perimeter_lf(self) -> float:
        return self.perimeter_lf_each * self.count

    @property
    def flashing_sqft_each(self) -> float:
        return self.perimeter_lf_each * (self.height_in / 12.0)

    @property
    def total_flashing_sqft(self) -> float:
        return self.flashing_sqft_each * self.count

    @property
    def labour_hours_per_curb(self) -> float:
        """Height-dependent labour (Excel: Takeoff I32-I36).
        Three height tiers applied to two rate tables (rip + install):
          < 25":  rip at 22.5 LF/hr, install at 15 LF/hr
          25-69": rip at 18 LF/hr,   install at 12 LF/hr
          > 69":  rip at 15 LF/hr,   install at 9 LF/hr
        """
        h = self.height_in
        perim = self.perimeter_lf_each
        if h < 25:
            return perim / 22.5 + perim / 15.0
        elif h <= 69:
            return perim / 18.0 + perim / 12.0
        else:
            return perim / 15.0 + perim / 9.0

    @property
    def total_labour_hours(self) -> float:
        return self.labour_hours_per_curb * self.count


@dataclass
class PerimeterSection:
    """One perimeter section A-E (Excel: Takeoff R52-R58).
    Each section has its own type, height, LF, and difficulty."""
    name: str = "A"
    perimeter_type: str = "parapet_no_facing"
    height_in: float = 24.0
    lf: float = 0.0
    fabrication_difficulty: str = "Normal"
    install_difficulty: str = "Normal"

    @property
    def strip_girth_in(self) -> float:
        """Membrane strip girth in inches (Excel: Takeoff G53-G57)."""
        h = self.height_in
        if self.perimeter_type == "parapet_no_facing":
            return h + 16   # 12" base run-out + 4" top
        elif self.perimeter_type == "parapet_w_facing":
            return h + 20   # 12" base + 8" top (facing overlap)
        elif self.perimeter_type == "interior_wall":
            return h + 16   # 12" base + 4" counter flash
        elif self.perimeter_type == "cant":
            return math.sqrt(2) * h + 12  # diagonal + 12" base
        elif self.perimeter_type == "divider_w_facing":
            return 2 * h + 12  # both sides + base
        return h + 16

    @property
    def strip_sqft(self) -> float:
        return (self.strip_girth_in / 12.0) * self.lf

    @property
    def metal_girth_in(self) -> float:
        """Metal flashing girth in inches (Excel: Takeoff H53-H57)."""
        h = self.height_in
        if self.perimeter_type == "parapet_no_facing":
            return h + 6    # 4" cap overlap + 2" hem
        elif self.perimeter_type == "parapet_w_facing":
            return 2 * h + 4  # cap covers both sides
        elif self.perimeter_type == "interior_wall":
            return h + 4
        elif self.perimeter_type == "cant":
            return 0         # no metal on cant
        elif self.perimeter_type == "divider_w_facing":
            return 2 * h + 8  # metal cap both sides
        return h + 6

    @property
    def metal_sqft(self) -> float:
        return (self.metal_girth_in / 12.0) * self.lf

    @property
    def metal_sheet_count(self) -> int:
        """Number of 10ft metal sheets needed."""
        if self.metal_girth_in == 0 or self.lf == 0:
            return 0
        return math.ceil(self.lf / 10.0)

    @property
    def top_of_parapet(self) -> bool:
        return self.perimeter_type in (
            "parapet_no_facing", "parapet_w_facing", "divider_w_facing"
        )

    @property
    def install_hours_per_sheet(self) -> float:
        rates = {"Easy": 0.5, "Normal": 0.75, "Hard": 1.0}
        return rates.get(self.install_difficulty, 0.75)

    @property
    def fabrication_hours_per_sheet(self) -> float:
        rates = {"Easy": 0.25, "Normal": 0.5, "Hard": 0.75}
        return rates.get(self.fabrication_difficulty, 0.5)

    @property
    def wood_face_sqft(self) -> float:
        """Wood facing area (only for types with facing)."""
        if self.perimeter_type in ("parapet_w_facing", "divider_w_facing"):
            return (self.height_in / 12.0) * self.lf
        return 0.0

    def install_hours(self, settings: ProjectSettings) -> float:
        """Install hours using Project sheet modifiers (Excel: Takeoff R53-R57)."""
        rate = settings.effective_rate
        if rate <= 0:
            return 0.0
        return (self.lf / 7.5) / rate


@dataclass
class VentItem:
    """Individual vent with type and difficulty (Excel: Takeoff R40-R50)."""
    vent_type: str = "pipe_boot"
    count: int = 0
    difficulty: str = "Normal"

    @property
    def hours_per_unit(self) -> float:
        info = VENT_LABOUR_HOURS.get(self.vent_type, {"base": 1.5})
        base = info["base"]
        adj = info.get(self.difficulty, 0.0)
        return base + adj

    @property
    def total_hours(self) -> float:
        return self.hours_per_unit * self.count


@dataclass
class WoodWorkSection:
    """Wood work section (Excel: Takeoff R67-R76)."""
    name: str = ""
    wood_type: str = "vertical"  # vertical, horizontal, plywood
    height_ft: float = 0.0
    lf: float = 0.0
    spacing_in: float = 16.0
    layers: int = 1
    lumber_size: str = "lumber_2x4"  # key into WOOD_PRODUCT_KEYS

    @property
    def quantity(self) -> float:
        """Number of pieces or sheets needed."""
        if self.wood_type == "plywood":
            return math.ceil((self.height_ft * self.lf) / 32.0) * self.layers
        elif self.wood_type == "vertical":
            if self.spacing_in <= 0:
                return 0
            return math.ceil(self.lf / (self.spacing_in / 12.0)) * self.layers
        else:  # horizontal
            if self.spacing_in <= 0:
                return 0
            rows = math.ceil(self.height_ft / (self.spacing_in / 12.0))
            return math.ceil(self.lf / 8.0) * rows * self.layers


@dataclass
class BattInsulationSection:
    """Batt insulation for pony walls (Excel: Takeoff R77-R83)."""
    name: str = ""
    height_ft: float = 0.0
    lf: float = 0.0
    insulation_type: str = "R24"
    layers: int = 1

    @property
    def sqft(self) -> float:
        return self.height_ft * self.lf * self.layers

    @property
    def bundles(self) -> int:
        return math.ceil(self.sqft / 40.0) if self.sqft > 0 else 0


# ---------------------------------------------------------------------------
# Unit conversion helpers (Excel: Takeoff J5-K11)
# ---------------------------------------------------------------------------

def mm_to_ft(mm: float) -> float:
    return mm / 304.8

def mm_to_in(mm: float) -> float:
    return mm / 25.4

def ft_to_mm(ft: float) -> float:
    return ft * 304.8

def in_to_mm(inches: float) -> float:
    return inches * 25.4


# ---------------------------------------------------------------------------
# Excel Takeoff Overrides (optional)
# ---------------------------------------------------------------------------

_TAKEOFF_CURB_ROW_MAP = {
    32: "RTU",
    33: "Roof_Hatch",
    34: "Vent_Curb",
    35: "Sleeper",
}

_TAKEOFF_VENT_ROW_MAP = {
    41: "pipe_boot",
    42: "b_vent",
    43: "hood_vent",
    44: "plumb_vent",
    45: "gum_box",
    46: "scupper",
    47: "radon_pipe",
    48: "drain",
}

_PERIMETER_TYPE_LOOKUP = {
    "parapet w/ facing": "parapet_w_facing",
    "parapet w/o facing": "parapet_no_facing",
    "interior walls": "interior_wall",
    "interior wall": "interior_wall",
    "cant": "cant",
    "divider w/ facing": "divider_w_facing",
}


def _normalize_text(value: str) -> str:
    return str(value or "").strip().lower()


def _normalize_difficulty(value: str | None, default: str = "Normal") -> str:
    if not value:
        return default
    cleaned = str(value).strip().replace(" ", "_").replace("-", "_")
    return cleaned


def load_takeoff_excel(path: str) -> dict:
    """Load curb/vent/perimeter inputs from the Excel Takeoff sheet."""
    if openpyxl is None:
        raise ImportError("openpyxl is required to load Excel takeoff data.")

    wb = openpyxl.load_workbook(path, data_only=True)
    if "Takeoff" not in wb.sheetnames:
        raise ValueError("Takeoff sheet not found in Excel workbook.")

    ws = wb["Takeoff"]

    curbs: list[CurbDetail] = []
    for row_idx, curb_type in _TAKEOFF_CURB_ROW_MAP.items():
        count = ws.cell(row=row_idx, column=3).value or 0
        length_ft = ws.cell(row=row_idx, column=4).value or 0
        width_ft = ws.cell(row=row_idx, column=5).value or 0
        height_in = ws.cell(row=row_idx, column=6).value or 0

        try:
            count = int(count)
        except (ValueError, TypeError):
            count = 0

        if count > 0:
            curbs.append(CurbDetail(
                curb_type=curb_type,
                count=count,
                length_in=float(length_ft) * 12.0,
                width_in=float(width_ft) * 12.0,
                height_in=float(height_in),
            ))

    vents: list[VentItem] = []
    for row_idx, vent_type in _TAKEOFF_VENT_ROW_MAP.items():
        count = ws.cell(row=row_idx, column=3).value or 0
        difficulty_raw = ws.cell(row=row_idx, column=4).value

        try:
            count = int(count)
        except (ValueError, TypeError):
            count = 0

        if count > 0:
            vents.append(VentItem(
                vent_type=vent_type,
                count=count,
                difficulty=_normalize_difficulty(difficulty_raw),
            ))

    perimeter_sections: list[PerimeterSection] = []
    for row_idx in range(53, 58):
        section_name = ws.cell(row=row_idx, column=2).value
        if not section_name:
            continue
        height_in = ws.cell(row=row_idx, column=3).value or 0
        type_raw = ws.cell(row=row_idx, column=5).value or ""
        lf = ws.cell(row=row_idx, column=6).value or 0
        fab_diff = ws.cell(row=row_idx, column=9).value or "Normal"
        install_diff = ws.cell(row=row_idx, column=10).value or "Normal"

        try:
            lf = float(lf)
        except (ValueError, TypeError):
            lf = 0.0

        if lf > 0:
            perimeter_sections.append(PerimeterSection(
                name=str(section_name),
                perimeter_type=_PERIMETER_TYPE_LOOKUP.get(
                    _normalize_text(type_raw), "parapet_no_facing"
                ),
                height_in=float(height_in or 0),
                lf=lf,
                fabrication_difficulty=str(fab_diff or "Normal"),
                install_difficulty=str(install_diff or "Normal"),
            ))

    corner_count = ws.cell(row=16, column=6).value
    try:
        corner_count = int(corner_count)
    except (ValueError, TypeError):
        corner_count = 0

    return {
        "curbs": curbs,
        "vents": vents,
        "perimeter_sections": perimeter_sections,
        "corner_count": corner_count,
    }


def apply_takeoff_excel(m: "RoofMeasurements", path: str) -> list[str]:
    """Apply Excel Takeoff overrides to a RoofMeasurements instance."""
    warnings: list[str] = []
    if not path:
        return warnings

    try:
        data = load_takeoff_excel(path)
    except Exception as exc:
        msg = f"Failed to load takeoff Excel '{path}': {exc}"
        logger.warning(msg)
        warnings.append(msg)
        return warnings

    if data.get("curbs"):
        m.curbs = data["curbs"]

    if data.get("vents"):
        m.vents = data["vents"]

    if data.get("perimeter_sections"):
        m.perimeter_sections = data["perimeter_sections"]
        total_lf = sum(s.lf for s in m.perimeter_sections)
        if total_lf > 0:
            m.perimeter_lf = total_lf
            top_lf = sum(s.lf for s in m.perimeter_sections if s.top_of_parapet)
            if top_lf > 0:
                m.parapet_length_lf = top_lf

    if data.get("corner_count"):
        m.corner_count = data["corner_count"]

    return warnings


def _aggregate_curbs(curbs: list[CurbDetail]) -> dict | None:
    if not curbs:
        return None
    total_count = sum(c.count for c in curbs if c.count > 0)
    if total_count <= 0:
        return None

    total_perimeter_lf = sum(c.total_perimeter_lf for c in curbs)
    total_flashing_sqft = sum(c.total_flashing_sqft for c in curbs)
    total_footprint_sqft = sum(
        (c.length_in / 12.0) * (c.width_in / 12.0) * c.count for c in curbs
    )
    avg_height_in = sum(c.height_in * c.count for c in curbs) / total_count

    return {
        "count": total_count,
        "perimeter_lf": total_perimeter_lf,
        "flashing_sqft": total_flashing_sqft,
        "footprint_sqft": total_footprint_sqft,
        "avg_height_in": avg_height_in,
    }


def _select_curb_group_by_area(curbs: list[CurbDetail], pick_largest: bool) -> list[CurbDetail]:
    groups: dict[str, list[CurbDetail]] = {}
    for curb in curbs:
        groups.setdefault(curb.curb_type, []).append(curb)

    if not groups:
        return []

    def area_for_group(group: list[CurbDetail]) -> float:
        # Assume consistent sizes per curb type; use first entry
        ref = group[0]
        return (ref.length_in / 12.0) * (ref.width_in / 12.0)

    selected = max(groups.values(), key=area_for_group) if pick_largest else min(
        groups.values(), key=area_for_group
    )
    return selected


def _detail_geometry_from_excel(detail_type: str, detail_name: str, m: "RoofMeasurements") -> dict | None:
    if not m.curbs:
        return None

    dtype = detail_type or ""
    name_lower = str(detail_name or "").lower()

    if dtype == "mechanical_curb":
        curbs = [c for c in m.curbs if c.curb_type == "RTU" and c.count > 0]
        if not curbs:
            curbs = [c for c in m.curbs if c.count > 0]
        return _aggregate_curbs(curbs)

    if dtype == "sleeper_curb":
        curbs = [c for c in m.curbs if c.curb_type == "Sleeper" and c.count > 0]
        return _aggregate_curbs(curbs)

    if dtype == "opening_cover":
        curbs = [c for c in m.curbs if c.count > 0]
        if not curbs:
            return None
        if "large" in name_lower:
            selected = _select_curb_group_by_area(curbs, pick_largest=True)
        elif "small" in name_lower:
            selected = _select_curb_group_by_area(curbs, pick_largest=False)
        else:
            for preferred in ("Vent_Curb", "Roof_Hatch", "RTU", "Sleeper"):
                selected = [c for c in curbs if c.curb_type == preferred and c.count > 0]
                if selected:
                    break
            else:
                selected = curbs
        return _aggregate_curbs(selected)

    return None


def _quantity_from_geometry(
    coverage: dict,
    geometry: dict | None,
    waste: float,
    dimension_in: float = 0.0,
) -> tuple[int, str] | None:
    if not geometry:
        return None
    unit = coverage.get("unit", "unit")

    if "lf_per_unit" in coverage and geometry.get("perimeter_lf"):
        lf_per = float(coverage.get("lf_per_unit") or 1)
        qty = math.ceil(geometry["perimeter_lf"] * waste / lf_per)
        return qty, unit

    if "sqft_per_unit" in coverage:
        sqft_per = float(coverage.get("sqft_per_unit") or 32)
        area = None
        if dimension_in > 0 and geometry.get("perimeter_lf"):
            area = geometry["perimeter_lf"] * (dimension_in / 12.0)
        elif geometry.get("flashing_sqft"):
            area = geometry["flashing_sqft"]
        elif geometry.get("footprint_sqft"):
            area = geometry["footprint_sqft"]

        if area and area > 0:
            qty = math.ceil(area * waste / sqft_per)
            return qty, unit

    return None

# ---------------------------------------------------------------------------
# System-Specific Material Definitions
# Coverage rates sourced from database.py descriptions
# ---------------------------------------------------------------------------

# Aggregated material sources for price lookup
_ALL_MATERIALS = [PRICING, EPDM_SPECIFIC_MATERIALS, TPO_SPECIFIC_MATERIALS, COMMON_ROOF_MATERIALS]

# Price overrides for materials with non-standard pricing models
_PRICE_OVERRIDES = {
    # EPS: $0.31/sqft/inch × 16 sqft (4'×4' sheet) × 2.5" thick = $12.40/sheet
    "EPS_Insulation_EPDM": 12.40,
}

# Area-based material layers per roof system type
# Format: (name, pricing_key, unit, sqft_per_unit, area_source, waste_pct, bid_group)
_SYSTEM_AREA_LAYERS = {
    "SBS": [
        ("Asphaltic Primer",
         "Primer", "pail (5 gal)", 250, "roof_area", 0.05, "roofing"),
        ("SBS Base Sheet (Sopraply Base 520)",
         "Base_Membrane", "roll", 100, "roof_area", 0.15, "roofing"),
        ("SBS Cap Sheet (Sopraply Traffic Cap)",
         "Cap_Membrane", "roll", 86, "roof_area", 0.15, "roofing"),
        ("Tapered ISO Insulation (Soprasmart Board 2:1)",
         "Polyisocyanurate_ISO_Insulation", "sheet (4'x4')", 16, "tapered_area", 0.10, "roofing"),
        ("XPS Insulation (Sopra-XPS 40 Type 4)",
         "XPS_Insulation", "sheet (2'x8')", 16, "roof_area", 0.10, "roofing"),
        ("Drainage Board (Sopradrain EcoVent)",
         "Drainage_Board", "roll (6'x50')", 300, "roof_area", 0.10, "roofing"),
        ("Filter Fabric",
         "Fleece_Reinforcement_Fabric", "roll", 300, "roof_area", 0.10, "roofing"),
    ],
    "EPDM_Fully_Adhered": [
        ("Vapour Barrier (Sopravap'r WG 45\")",
         "Vapour_Barrier_Sopravapor", "roll (45\" x 5Sq)", 500, "roof_area", 0.10, "roofing"),
        ("ISO Insulation 2.5\" (Sopra-ISO)",
         "ISO_2_5_inch", "sheet (4'x4')", 16, "roof_area", 0.10, "roofing"),
        ("Tapered ISO Insulation (drainage slope)",
         "Tapered_ISO", "sqft", 1, "tapered_area", 0.10, "roofing"),
        ("Densdeck Coverboard 1/2\"",
         "Densdeck_Half_Inch", "sheet (4'x8')", 32, "roof_area", 0.10, "roofing"),
        ("EPDM Membrane 60 mil (Carlisle Sure-Seal)",
         "EPDM_Membrane_60mil", "roll (10'x100')", 1000, "roof_area", 0.10, "roofing"),
        ("EPDM Bonding Adhesive 90-8-30A",
         "EPDM_Bonding_Adhesive", "pail (5 gal)", 300, "roof_area", 0.05, "roofing"),
        ("EPDM Primer HP-250",
         "EPDM_Primer_HP250", "gallon", 50, "roof_area", 0.05, "roofing"),
        ("EPDM Seam Tape 3\"x100'",
         "EPDM_Seam_Tape", "roll (100 lf)", 1000, "roof_area", 0.10, "roofing"),
    ],
    "EPDM_Ballasted": [
        ("Vapour Barrier (Sopravap'r WG 45\")",
         "Vapour_Barrier_Sopravapor", "roll (45\" x 5Sq)", 500, "roof_area", 0.10, "roofing"),
        ("EPDM Membrane 60 mil (Carlisle Sure-Seal, loose laid)",
         "EPDM_Membrane_60mil", "roll (10'x100')", 1000, "roof_area", 0.10, "roofing"),
        ("EPS Insulation Type II (2 layers x 2.5\")",
         "EPS_Insulation_EPDM", "sheet (4'x4')", 8, "roof_area", 0.10, "roofing"),
        ("Filter Fabric (Soprafilter)",
         "EPDM_Filter_Fabric", "roll", 300, "roof_area", 0.10, "roofing"),
        ("Drainage Mat (Sopradrain 15G 6'x50')",
         "EPDM_Drainage_Mat", "roll (6'x50')", 300, "roof_area", 0.10, "roofing"),
        ("EPDM Seam Tape 3\"x100'",
         "EPDM_Seam_Tape", "roll (100 lf)", 1000, "roof_area", 0.10, "roofing"),
    ],
    "TPO_Mechanically_Attached": [
        ("Vapour Barrier (Sopravap'r WG 45\")",
         "Vapour_Barrier_Sopravapor", "roll (45\" x 5Sq)", 500, "roof_area", 0.10, "roofing"),
        ("ISO Insulation 2.5\" (Sopra-ISO)",
         "ISO_2_5_inch", "sheet (4'x4')", 16, "roof_area", 0.10, "roofing"),
        ("Tapered ISO Insulation (drainage slope)",
         "Tapered_ISO", "sqft", 1, "tapered_area", 0.10, "roofing"),
        ("Densdeck Coverboard 1/2\"",
         "Densdeck_Half_Inch", "sheet (4'x8')", 32, "roof_area", 0.10, "roofing"),
        ("TPO Membrane 60 mil (Sure-Weld)",
         "TPO_Membrane", "roll (10'x100')", 1000, "roof_area", 0.10, "roofing"),
        ("Rhinobond Induction Weld Plates",
         "TPO_Rhinobond_Plate", "pallet", 4000, "roof_area", 0.05, "roofing"),
        ("TPO Fastening Screws",
         "TPO_Screws", "box", 4000, "roof_area", 0.05, "roofing"),
    ],
    "TPO_Fully_Adhered": [
        ("Vapour Barrier (Sopravap'r WG 45\")",
         "Vapour_Barrier_Sopravapor", "roll (45\" x 5Sq)", 500, "roof_area", 0.10, "roofing"),
        ("ISO Insulation 2.5\" (Sopra-ISO)",
         "ISO_2_5_inch", "sheet (4'x4')", 16, "roof_area", 0.10, "roofing"),
        ("Tapered ISO Insulation (drainage slope)",
         "Tapered_ISO", "sqft", 1, "tapered_area", 0.10, "roofing"),
        ("Soprasmart ISO HD 1/2\" (factory laminated coverboard)",
         "Soprasmart_ISO_HD", "sheet (4'x8')", 32, "roof_area", 0.10, "roofing"),
        ("TPO Membrane 60 mil (Sure-Weld)",
         "TPO_Membrane", "roll (10'x100')", 1000, "roof_area", 0.10, "roofing"),
        ("TPO Bonding Adhesive (SureWeld)",
         "TPO_Bonding_Adhesive_SureWeld", "pail (5 gal)", 300, "roof_area", 0.05, "roofing"),
        ("TPO Primer",
         "TPO_Primer", "gallon", 100, "roof_area", 0.05, "roofing"),
    ],
}

# Consumables per system type
# Format: (name, pricing_key, unit, rate_per_1000sqft, bid_group)
_SYSTEM_CONSUMABLES = {
    "SBS": [
        # Adhesives - wall vs field split (Excel: FRS R41-R46)
        ("Mastic (Sopramastic)", "Mastic", "pail", 2, "roofing"),
        ("Elastocol Adhesive - Field", "Adhesive_Elastocol", "pail (19L)", 3, "roofing"),
        ("Sealant (Mulco Supra)", "Sealant_Mulco_Supra", "tube", 6, "flashing"),
        # Accessories (Excel: FRS R48-R59) — Firetape moved to conditional calc
        ("Sopralap Cover Strip", "Sopralap_Cover_Strip", "roll", 1, "roofing"),
        ("Screws & Plates (insulation fastening)", "Screws_Plates_Combo", "box (1M)", 1, "roofing"),
        ("Flashing Bond Mastic", "Flashing_Bond_Mastic", "tube", 3, "flashing"),
    ],
    "EPDM_Fully_Adhered": [
        ("EPDM Lap Sealant", "EPDM_Lap_Sealant", "tube", 4, "roofing"),
        ("Duotack Foamable Adhesive (insulation bonding)", "Duotack_Adhesive", "case", 2, "roofing"),
        ("Polyurethane Sealant (Dymonic 100 / NP1)", "Sealant_General", "tube", 4, "flashing"),
        ("EPDM Cav Grip Adhesive", "EPDM_Cav_Grip", "cylinder", 0.5, "roofing"),
        ("Screws & Plates (insulation fastening)", "Screws_Plates_Combo", "box (1M)", 1, "roofing"),
    ],
    "EPDM_Ballasted": [
        ("EPDM Lap Sealant", "EPDM_Lap_Sealant", "tube", 4, "roofing"),
        ("Polyurethane Sealant (Dymonic 100 / NP1)", "Sealant_General", "tube", 4, "flashing"),
    ],
    "TPO_Mechanically_Attached": [
        ("TPO Lap Sealant", "TPO_Lap_Sealant", "tube", 3, "roofing"),
        ("TPO Tuck Tape", "TPO_Tuck_Tape", "roll", 2, "roofing"),
        ("Polyurethane Sealant (Dymonic 100 / NP1)", "Sealant_General", "tube", 4, "flashing"),
        ("Screws & Plates (insulation fastening)", "Screws_Plates_Combo", "box (1M)", 1, "roofing"),
    ],
    "TPO_Fully_Adhered": [
        ("TPO Lap Sealant", "TPO_Lap_Sealant", "tube", 3, "roofing"),
        ("TPO Tuck Tape", "TPO_Tuck_Tape", "roll", 2, "roofing"),
        ("Polyurethane Sealant (Dymonic 100 / NP1)", "Sealant_General", "tube", 4, "flashing"),
    ],
}

# Wall-only consumables: computed from parapet strip sqft, not roof area
# Format: (name, pricing_key, unit, sqft_per_unit, bid_group)
_WALL_CONSUMABLES = {
    "SBS": [
        ("Elastocol Adhesive - Wall (parapet strips)", "Adhesive_Elastocol", "pail (19L)", 333, "flashing"),
    ],
    "EPDM_Fully_Adhered": [
        ("EPDM Primer HP-250 (wall details)", "EPDM_Primer_HP250", "gallon", 50, "flashing"),
    ],
    "EPDM_Ballasted": [],
    "TPO_Mechanically_Attached": [
        ("TPO Primer (wall details)", "TPO_Primer", "gallon", 100, "flashing"),
    ],
    "TPO_Fully_Adhered": [
        ("TPO Primer (wall details)", "TPO_Primer", "gallon", 100, "flashing"),
    ],
}

# System metadata for display and bid multipliers
_SYSTEM_META = {
    "SBS": {
        "display_name": "Inverted Modified Bitumen (2-Ply SBS) - Soprema System",
        "spec": "Div 07 52 01 / 07 62 00 / 07 92 00",
        "labour_multiplier": 1.65,
        "labour_note": "Labour typically 1.5-1.8x material for SBS torch-applied",
        "mechanical_multiplier": 1.80,
        "include_ballast_note": True,
    },
    "EPDM_Fully_Adhered": {
        "display_name": "EPDM 60 mil Fully Adhered System",
        "spec": "Div 07 53 23",
        "labour_multiplier": 1.55,
        "labour_note": "Labour typically 1.4-1.6x material for EPDM fully adhered",
        "mechanical_multiplier": 1.80,
        "include_ballast_note": False,
    },
    "EPDM_Ballasted": {
        "display_name": "EPDM 60 mil Ballasted / Inverted System",
        "spec": "Div 07 53 23",
        "labour_multiplier": 1.40,
        "labour_note": "Labour typically 1.3-1.5x material for EPDM ballasted",
        "mechanical_multiplier": 1.70,
        "include_ballast_note": True,
    },
    "TPO_Mechanically_Attached": {
        "display_name": "TPO 60 mil Mechanically Attached System",
        "spec": "Div 07 54 23",
        "labour_multiplier": 1.50,
        "labour_note": "Labour typically 1.4-1.6x material for TPO mechanically attached",
        "mechanical_multiplier": 1.80,
        "include_ballast_note": False,
    },
    "TPO_Fully_Adhered": {
        "display_name": "TPO 60 mil Fully Adhered System",
        "spec": "Div 07 54 23",
        "labour_multiplier": 1.65,
        "labour_note": "Labour typically 1.5-1.8x material for TPO fully adhered",
        "mechanical_multiplier": 1.80,
        "include_ballast_note": False,
    },
}

# System-specific pipe seal product keys
_PIPE_SEAL_KEY = {
    "SBS": ("Pipe_Boot_Seal", "Penetration Seal"),
    "EPDM_Fully_Adhered": ("EPDM_Pipe_Flashing", "EPDM Pipe Flashing (1\"-6\")"),
    "EPDM_Ballasted": ("EPDM_Pipe_Flashing", "EPDM Pipe Flashing (1\"-6\")"),
    "TPO_Mechanically_Attached": ("TPO_Pipe_Boot", "TPO Universal Pipe Boot"),
    "TPO_Fully_Adhered": ("TPO_Pipe_Boot", "TPO Universal Pipe Boot"),
}


# ---------------------------------------------------------------------------
# Project Measurements (input from scaled drawings)
# ---------------------------------------------------------------------------

@dataclass
class RoofMeasurements:
    """Measurements taken from architectural drawings (plan + section views)."""

    # --- Field area (from plan view) ---
    total_roof_area_sqft: float
    perimeter_lf: float

    # --- Parapet (simple — used when perimeter_sections is empty) ---
    parapet_length_lf: float
    parapet_height_ft: float = 2.0

    # --- Multi-section flat roof (Excel: Takeoff R5-R14) ---
    roof_sections: list = field(default_factory=list)  # list[RoofSection]

    # --- Unit conversion (Excel: Takeoff J5-K11) ---
    input_unit: str = "imperial"  # "imperial" or "metric"

    # --- Curbs with dimensions (Excel: Takeoff R31-R37) ---
    curbs: list = field(default_factory=list)  # list[CurbDetail]
    extra_mechanical_hours: float = 0.0

    # --- Perimeter sections A-E (Excel: Takeoff R52-R58) ---
    perimeter_sections: list = field(default_factory=list)  # list[PerimeterSection]
    corner_count: int = 0

    # --- Vents with types (Excel: Takeoff R40-R50) ---
    vents: list = field(default_factory=list)  # list[VentItem]

    # --- Legacy simple counts (backward compat — used when vents/curbs empty) ---
    roof_drain_count: int = 0
    scupper_count: int = 0
    mechanical_unit_count: int = 0
    sleeper_curb_count: int = 0
    vent_hood_count: int = 0
    gas_penetration_count: int = 0
    electrical_penetration_count: int = 0
    plumbing_vent_count: int = 0
    gum_box_count: int = 0
    b_vent_count: int = 0
    radon_pipe_count: int = 0
    roof_hatch_count: int = 0
    skylight_count: int = 0

    # --- Optional area overrides ---
    tapered_area_sqft: float | None = None
    ballast_area_sqft: float | None = None

    # --- System type ---
    roof_system_type: str = "SBS"

    # --- Wood work (Excel: Takeoff R67-R76) ---
    wood_sections: list = field(default_factory=list)  # list[WoodWorkSection]

    # --- Batt insulation (Excel: Takeoff R77-R83) ---
    batt_sections: list = field(default_factory=list)  # list[BattInsulationSection]

    # --- Other costs (Excel: FRS R123-R125) ---
    delivery_count: int = 1
    disposal_roof_count: int = 1
    include_toilet: bool = False
    include_fencing: bool = False

    # --- Metal flashing type selection ---
    metal_flashing_type: str = "galvanized"

    # --- Material enable/disable toggles (Excel: FRS D column) ---
    include_vapour_barrier: bool = True
    include_insulation: bool = True
    include_coverboard: bool = True
    include_tapered: bool = True
    include_drainage: bool = True

    # --- Vapour barrier tie-in (Excel: FRS R18) ---
    vapour_barrier_tie_in: bool = False

    # --- SBS base sheet type ---
    sbs_base_type: str = "torch"  # "torch" or "peel_stick"

    # --- Project settings (Excel: Project sheet modifiers) ---
    project_settings: ProjectSettings = field(default_factory=ProjectSettings)

    # --- Fire Prevention Board (Excel: FRS R29) ---
    fire_board_scope: str = "None"  # "None", "Wall", "Field", "Both"

    # --- Vapour barrier details (Excel: FRS R17, for firetape conditional) ---
    vapour_barrier_attachment: str = "Torched"  # "Mopped", "Torched", "Self-Adhered"
    vapour_barrier_product: str = "Sopravapor"  # "#15_Felt_x2", "Sopravapor"

    # --- Asphalt EasyMelt (Excel: FRS R46) ---
    include_asphalt_easymelt: bool = False

    # --- PMMA system (Excel: FRS D31) ---
    include_pmma: bool = False

    # --- Garland system (Excel: FRS R56-R59) ---
    garland_system: bool = False

    # --- Optional ISO layers (Excel: FRS R22-R23) ---
    second_iso_layer: bool = False
    third_iso_layer: bool = False

    # --- Version (Excel: Takeoff B3) ---
    version: str = ""

    # --- Ballast gravel type (Excel: FRS R44) ---
    ballast_type: str = "BUR"  # "BUR" or "EPDM"

    # --- EPS thickness (Excel: FRS A64) ---
    eps_thickness_in: float = 2.5

    # --- TPO toggles (Excel: FRS D88-D91) ---
    tpo_second_membrane: bool = False
    include_tpo_flashing_24: bool = True
    include_tpo_flashing_12: bool = False

    # ---------------------------------------------------------------
    # Computed properties
    # ---------------------------------------------------------------

    @property
    def computed_roof_area(self) -> float:
        """Use multi-section areas if provided, else fallback to total."""
        if self.roof_sections:
            return sum(s.area_sqft for s in self.roof_sections)
        return self.total_roof_area_sqft

    @property
    def computed_parapet_lf(self) -> float:
        """Use perimeter sections if provided, else fallback."""
        if self.perimeter_sections:
            return sum(s.lf for s in self.perimeter_sections)
        return self.parapet_length_lf

    @property
    def total_strip_sqft(self) -> float:
        """Total membrane strip area from perimeter sections."""
        if self.perimeter_sections:
            return sum(s.strip_sqft for s in self.perimeter_sections)
        return self.parapet_length_lf * self.parapet_height_ft

    @property
    def total_metal_sqft(self) -> float:
        """Total metal flashing area from perimeter sections."""
        if self.perimeter_sections:
            return sum(s.metal_sqft for s in self.perimeter_sections)
        return self.parapet_length_lf * (self.parapet_height_ft + 0.5)

    @property
    def total_metal_sheets(self) -> int:
        """Total metal sheets from perimeter sections."""
        if self.perimeter_sections:
            return sum(s.metal_sheet_count for s in self.perimeter_sections)
        return math.ceil(self.parapet_length_lf / 10.0) if self.parapet_length_lf > 0 else 0

    @property
    def total_wood_face_sqft(self) -> float:
        """Total wood facing area from perimeter sections (facing types only)."""
        return sum(s.wood_face_sqft for s in self.perimeter_sections)

    @property
    def total_curb_perimeter_lf(self) -> float:
        return sum(c.total_perimeter_lf for c in self.curbs)

    @property
    def total_curb_flashing_sqft(self) -> float:
        return sum(c.total_flashing_sqft for c in self.curbs)

    @property
    def total_curb_labour_hours(self) -> float:
        return sum(c.total_labour_hours for c in self.curbs) + self.extra_mechanical_hours

    @property
    def total_perimeter_install_hours(self) -> float:
        """Total perimeter install hours using project settings."""
        return sum(s.install_hours(self.project_settings) for s in self.perimeter_sections)

    @property
    def total_vent_hours(self) -> float:
        return sum(v.total_hours for v in self.vents)

    @property
    def total_vent_count(self) -> int:
        if self.vents:
            return sum(v.count for v in self.vents)
        return (self.roof_drain_count + self.scupper_count +
                self.vent_hood_count + self.gas_penetration_count +
                self.electrical_penetration_count + self.plumbing_vent_count +
                self.gum_box_count + self.b_vent_count + self.radon_pipe_count)

    @property
    def effective_tapered_area(self) -> float:
        area = self.computed_roof_area
        return self.tapered_area_sqft if self.tapered_area_sqft is not None else area

    @property
    def effective_ballast_area(self) -> float:
        area = self.computed_roof_area
        return self.ballast_area_sqft if self.ballast_area_sqft is not None else area

    @property
    def total_penetrations(self) -> int:
        if self.curbs or self.vents:
            curb_total = sum(c.count for c in self.curbs)
            vent_total = sum(v.count for v in self.vents)
            return curb_total + vent_total + self.roof_hatch_count + self.skylight_count
        return (self.mechanical_unit_count + self.sleeper_curb_count +
                self.vent_hood_count + self.gas_penetration_count +
                self.electrical_penetration_count + self.plumbing_vent_count)


def validate_measurements(m: RoofMeasurements) -> list[str]:
    """
    Validate measurements and return a list of warning messages.
    Does not block execution, just flags suspicious values.
    """
    warnings = []
    
    if m.total_roof_area_sqft <= 0:
        warnings.append("Total roof area is zero or negative.")
    
    if m.perimeter_lf <= 0:
        warnings.append("Roof perimeter is zero or negative.")
        
    if m.total_roof_area_sqft > 0 and m.perimeter_lf > 0:
        # Check for unreasonable area/perimeter ratio (e.g. extremely long/thin or error)
        # A square has P = 4 * sqrt(A). If P is vastly smaller, it's physically impossible.
        min_perimeter = 4 * math.sqrt(m.total_roof_area_sqft)
        if m.perimeter_lf < min_perimeter * 0.5: # Allow some margin for error/shape
            warnings.append(f"Perimeter ({m.perimeter_lf:.0f}') seems too small for the area ({m.total_roof_area_sqft:.0f} sqft).")

    if m.parapet_length_lf > m.perimeter_lf * 1.5:
        warnings.append("Parapet length is significantly longer than roof perimeter.")
        
    if m.parapet_height_ft > 6.0:
        warnings.append(f"Parapet height ({m.parapet_height_ft} ft) is unusually high.")
        
    return warnings


# ---------------------------------------------------------------------------
# Assembly Definition - Inverted Modified Bitumen (2-Ply SBS, Soprema)
# Spec: Div 07 52 01 / 07 62 00 / 07 92 00
# ---------------------------------------------------------------------------

def _get_price(pricing_key: str) -> float:
    """Look up avg_price from any material dictionary. Returns 0 if key missing."""
    if pricing_key in _PRICE_OVERRIDES:
        return _PRICE_OVERRIDES[pricing_key]
    for source in _ALL_MATERIALS:
        entry = source.get(pricing_key)
        if entry is not None:
            if isinstance(entry, dict):
                return entry.get("avg_price", 0.0)
            return float(entry)
    return 0.0


def calculate_takeoff(m: RoofMeasurements) -> dict:
    """
    Calculate full material quantity takeoff and cost estimate.

    Covers Takeoff Sheet + Flat Roof Summary from SBS_Worksheet_4_5.xlsm:
      - Multi-section roof areas
      - Perimeter girth calculations (6 types)
      - Curb dimensioned calculations
      - Vent type/difficulty hours
      - EPDM/TPO specific quantity formulas
      - Wood work and batt insulation
      - Wall vs field adhesive split
      - Material toggles
      - Other costs (delivery, disposal, toilet, fencing)
    """

    system = m.roof_system_type
    meta = _SYSTEM_META.get(system, _SYSTEM_META["SBS"])
    roof_area = m.computed_roof_area
    parapet_lf = m.computed_parapet_lf

    results = {
        "project_measurements": {
            "total_roof_area_sqft": roof_area,
            "perimeter_lf": m.perimeter_lf,
            "parapet_length_lf": parapet_lf,
            "parapet_height_ft": m.parapet_height_ft,
            "tapered_area_sqft": m.effective_tapered_area,
            "ballast_area_sqft": m.effective_ballast_area,
            "total_penetrations": m.total_penetrations,
            "corner_count": m.corner_count,
            "roof_system_type": system,
            "roof_system_name": meta["display_name"],
            "spec": meta["spec"],
            "layer_count": len(_SYSTEM_AREA_LAYERS.get(system, [])),
            "version": m.version,
            "total_curb_labour_hours": round(m.total_curb_labour_hours, 1),
            "total_perimeter_install_hours": round(m.total_perimeter_install_hours, 1),
            "total_vent_hours": round(m.total_vent_hours, 1),
        },
        "roof_sections": [],
        "area_materials": [],
        "perimeter_details": [],
        "linear_materials": [],
        "curb_details": [],
        "vent_details": [],
        "unit_items": [],
        "consumables": [],
        "epdm_tpo_details": [],
        "wood_materials": [],
        "batt_insulation": [],
        "other_costs": [],
    }

    total_roofing_cost = 0.0
    total_flashing_cost = 0.0
    total_mechanical_cost = 0.0
    total_other_cost = 0.0

    # ===================================================================
    # MULTI-SECTION ROOF AREAS (Excel: Takeoff R5-R14)
    # ===================================================================
    if m.roof_sections:
        for sec in m.roof_sections:
            if sec.area_sqft > 0:
                results["roof_sections"].append({
                    "name": sec.name,
                    "count": sec.count,
                    "length_ft": sec.length_ft,
                    "width_ft": sec.width_ft,
                    "area_sqft": round(sec.area_sqft, 0),
                })
        results["project_measurements"]["total_roof_area_sqft"] = roof_area

    # ===================================================================
    # AREA-BASED MATERIALS (membrane, insulation, drainage)
    # With material toggle support (Excel: FRS D column Yes/No)
    # ===================================================================
    area_layers = _SYSTEM_AREA_LAYERS.get(system, _SYSTEM_AREA_LAYERS["SBS"])

    # Map pricing keys to toggle categories
    _TOGGLE_MAP = {
        "Vapour_Barrier_Sopravapor": "include_vapour_barrier",
        "Polyisocyanurate_ISO_Insulation": "include_insulation",
        "ISO_2_5_inch": "include_insulation",
        "XPS_Insulation": "include_insulation",
        "EPS_Insulation_EPDM": "include_insulation",
        "DensDeck_Coverboard": "include_coverboard",
        "Densdeck_Half_Inch": "include_coverboard",
        "Soprasmart_ISO_HD": "include_coverboard",
        "Tapered_ISO": "include_tapered",
        "Drainage_Board": "include_drainage",
        "EPDM_Drainage_Mat": "include_drainage",
        "EPDM_Filter_Fabric": "include_drainage",
        "Fleece_Reinforcement_Fabric": "include_drainage",
    }

    for name, pkey, unit, sqft_per_unit, area_src, waste_pct, bid_grp in area_layers:
        # Check material toggle
        toggle_attr = _TOGGLE_MAP.get(pkey)
        if toggle_attr and not getattr(m, toggle_attr, True):
            continue

        if area_src == "roof_area":
            base_area = roof_area
        elif area_src == "tapered_area":
            base_area = m.effective_tapered_area
        elif area_src == "ballast_area":
            base_area = m.effective_ballast_area
        else:
            base_area = roof_area

        area_with_waste = base_area * (1 + waste_pct)
        qty = math.ceil(area_with_waste / sqft_per_unit)
        # EPS thickness-tiered pricing: $0.31/sqft/inch × sheet area × thickness
        if pkey == "EPS_Insulation_EPDM":
            unit_price = 0.31 * m.eps_thickness_in * 16  # 4'×4' = 16 sqft
        else:
            unit_price = _get_price(pkey)
        line_cost = qty * unit_price

        results["area_materials"].append({
            "name": name,
            "base_area_sqft": round(base_area, 0),
            "waste_pct": f"{waste_pct:.0%}",
            "quantity": qty,
            "unit": unit,
            "unit_price": round(unit_price, 2),
            "line_cost": round(line_cost, 2),
            "bid_group": bid_grp,
        })

        if bid_grp == "roofing":
            total_roofing_cost += line_cost

    # Gravel ballast (Excel: FRS R44) — BUR vs EPDM type
    if meta["include_ballast_note"]:
        squares = roof_area / 100.0
        if m.ballast_type == "EPDM":
            ballast_qty = math.ceil(squares / 3)
            ballast_label = "EPDM Gravel Ballast"
        else:
            ballast_qty = math.ceil(squares / 6)
            ballast_label = "BUR Gravel Ballast"
        results["area_materials"].append({
            "name": f"{ballast_label} (redistribute existing)",
            "base_area_sqft": round(m.effective_ballast_area, 0),
            "waste_pct": "0%",
            "quantity": ballast_qty,
            "unit": "loads",
            "unit_price": 0.0,
            "line_cost": 0.0,
            "bid_group": "roofing",
            "note": "Existing ballast. Reduce depth to 100mm max.",
        })

    # Vapour Barrier Tie-In (Excel: FRS R18)
    if m.vapour_barrier_tie_in:
        vb_price = _get_price("Vapour_Barrier_TieIn")
        results["area_materials"].append({
            "name": "Vapour Barrier Tie-In Allowance",
            "base_area_sqft": 0,
            "waste_pct": "0%",
            "quantity": 1,
            "unit": "allowance",
            "unit_price": round(vb_price, 2),
            "line_cost": round(vb_price, 2),
            "bid_group": "roofing",
        })
        total_roofing_cost += vb_price

    # Fire Prevention Board (Excel: FRS R29)
    if m.fire_board_scope != "None":
        wall_area = m.total_strip_sqft
        wall_fb_qty = math.ceil(wall_area / 20 * 1.1) if wall_area > 0 else 0
        field_fb_qty = math.ceil(roof_area / 20 * 1.1)
        if m.fire_board_scope == "Wall":
            fb_qty = wall_fb_qty
        elif m.fire_board_scope == "Field":
            fb_qty = field_fb_qty
        else:  # "Both"
            fb_qty = wall_fb_qty + field_fb_qty
        fb_price = _get_price("Fire_Prevention_Board")
        fb_cost = fb_qty * fb_price
        results["area_materials"].append({
            "name": f"Fire Prevention Board ({m.fire_board_scope})",
            "base_area_sqft": round(wall_area + roof_area if m.fire_board_scope == "Both"
                                    else wall_area if m.fire_board_scope == "Wall"
                                    else roof_area, 0),
            "waste_pct": "10%",
            "quantity": fb_qty,
            "unit": "sheet (20 sqft)",
            "unit_price": round(fb_price, 2),
            "line_cost": round(fb_cost, 2),
            "bid_group": "roofing",
        })
        total_roofing_cost += fb_cost

    # Optional 2nd/3rd ISO Insulation layers (Excel: FRS R22-R23)
    for layer_num, enabled in [(2, m.second_iso_layer), (3, m.third_iso_layer)]:
        if enabled and m.include_insulation:
            iso_qty = math.ceil(roof_area * 1.1 / 16)
            iso_price = _get_price("ISO_2_5_inch")
            iso_cost = iso_qty * iso_price
            results["area_materials"].append({
                "name": f"ISO Insulation 2.5\" - Layer {layer_num}",
                "base_area_sqft": round(roof_area, 0),
                "waste_pct": "10%",
                "quantity": iso_qty,
                "unit": "sheet (4'x4')",
                "unit_price": round(iso_price, 2),
                "line_cost": round(iso_cost, 2),
                "bid_group": "roofing",
            })
            total_roofing_cost += iso_cost

    # ===================================================================
    # PERIMETER SECTION DETAILS (Excel: Takeoff R52-R58)
    # Girth calculations per section type
    # ===================================================================
    if m.perimeter_sections:
        for sec in m.perimeter_sections:
            if sec.lf <= 0:
                continue
            results["perimeter_details"].append({
                "name": sec.name,
                "type": PERIMETER_TYPES.get(sec.perimeter_type, sec.perimeter_type),
                "height_in": sec.height_in,
                "lf": round(sec.lf, 0),
                "strip_girth_in": round(sec.strip_girth_in, 1),
                "strip_sqft": round(sec.strip_sqft, 0),
                "metal_girth_in": round(sec.metal_girth_in, 1),
                "metal_sqft": round(sec.metal_sqft, 0),
                "metal_sheets": sec.metal_sheet_count,
                "top_of_parapet": sec.top_of_parapet,
                "wood_face_sqft": round(sec.wood_face_sqft, 0),
                "fab_difficulty": sec.fabrication_difficulty,
                "install_difficulty": sec.install_difficulty,
                "install_hours": round(sec.install_hours(m.project_settings), 1),
            })

    # ===================================================================
    # LINEAR-FOOT MATERIALS (flashings, blocking, sheathing)
    # Uses perimeter section girth data when available
    # ===================================================================
    metal_flash_key = METAL_FLASHING_TYPES.get(
        m.metal_flashing_type, "Metal_Flashing_Galvanized"
    )
    metal_type_label = {
        "galvanized": "Galvanized w/ Clips",
        "prepainted": "Prepainted",
        "cladding": "Cladding Panel",
    }.get(m.metal_flashing_type, "Galvanized")

    if m.perimeter_sections:
        # Girth-based calculation: metal from perimeter section data
        total_cap_lf = sum(
            s.lf for s in m.perimeter_sections if s.top_of_parapet and s.lf > 0
        )
        total_counter_lf = sum(s.lf for s in m.perimeter_sections if s.lf > 0)
        total_wood_lf = parapet_lf
        total_ply_lf = parapet_lf

        linear_items = [
            (f"Metal Cap Flashing ({metal_type_label})",
             metal_flash_key, "LF", 1, total_cap_lf, 0.10, "flashing"),
            (f"Metal Counter Flashing ({metal_type_label})",
             metal_flash_key, "LF", 1, total_counter_lf, 0.10, "flashing"),
            ("Wood Blocking (SPF 2x)",
             "Wood_Blocking_Lumber", "8ft piece", 8, total_wood_lf, 0.15, "flashing"),
            ("Plywood Sheathing (12.5mm Douglas Fir)",
             "Plywood_Sheathing", "4'x8' sheet", 8, total_ply_lf, 0.15, "flashing"),
        ]
    else:
        # Simple fallback
        linear_items = [
            (f"Metal Cap Flashing ({metal_type_label})",
             metal_flash_key, "LF", 1, m.parapet_length_lf, 0.10, "flashing"),
            (f"Metal Counter Flashing ({metal_type_label})",
             metal_flash_key, "LF", 1, m.parapet_length_lf, 0.10, "flashing"),
            ("Wood Blocking (SPF 2x)",
             "Wood_Blocking_Lumber", "8ft piece", 8, m.parapet_length_lf, 0.15, "flashing"),
            ("Plywood Sheathing (12.5mm Douglas Fir)",
             "Plywood_Sheathing", "4'x8' sheet", 8, m.parapet_length_lf, 0.15, "flashing"),
        ]

    for name, pkey, unit, lf_per_unit, base_lf, waste_pct, bid_grp in linear_items:
        if base_lf <= 0:
            continue
        lf_with_waste = base_lf * (1 + waste_pct)
        qty = math.ceil(lf_with_waste / lf_per_unit)
        unit_price = _get_price(pkey)
        line_cost = qty * unit_price

        results["linear_materials"].append({
            "name": name,
            "base_lf": round(base_lf, 0),
            "waste_pct": f"{waste_pct:.0%}",
            "quantity": qty,
            "unit": unit,
            "unit_price": round(unit_price, 2),
            "line_cost": round(line_cost, 2),
            "bid_group": bid_grp,
        })

        if bid_grp == "flashing":
            total_flashing_cost += line_cost

    # ===================================================================
    # CURB DETAILS (Excel: Takeoff R31-R37)
    # Dimensioned curbs with perimeter, flashing sqft, labour hours
    # ===================================================================
    if m.curbs:
        for curb in m.curbs:
            if curb.count <= 0:
                continue
            # Curb flashing material
            flash_sqft = curb.total_flashing_sqft
            flash_price = _get_price("Flashing_General")
            # Price per sqft of flashing (estimate: each 10ft piece covers ~3.33 sqft at 4" girth)
            flash_lf = curb.total_perimeter_lf
            flash_pcs = math.ceil(flash_lf * 1.1 / 10.0)
            flash_cost = flash_pcs * flash_price

            results["curb_details"].append({
                "curb_type": curb.curb_type,
                "count": curb.count,
                "dimensions": f"{curb.length_in}\"L x {curb.width_in}\"W x {curb.height_in}\"H",
                "perimeter_lf": round(curb.total_perimeter_lf, 1),
                "flashing_sqft": round(flash_sqft, 0),
                "labour_hours": round(curb.total_labour_hours, 1),
                "flashing_pieces": flash_pcs,
                "flashing_cost": round(flash_cost, 2),
            })
            total_mechanical_cost += flash_cost

    # Extra mechanical hours
    if m.extra_mechanical_hours > 0:
        results["curb_details"].append({
            "curb_type": "Extra Mechanical Hours",
            "count": 1,
            "dimensions": "-",
            "perimeter_lf": 0,
            "flashing_sqft": 0,
            "labour_hours": m.extra_mechanical_hours,
            "flashing_pieces": 0,
            "flashing_cost": 0,
        })

    # ===================================================================
    # VENT DETAILS (Excel: Takeoff R40-R50)
    # Type-specific with difficulty hour adjustments
    # ===================================================================
    if m.vents:
        for vent in m.vents:
            if vent.count <= 0:
                continue
            results["vent_details"].append({
                "vent_type": vent.vent_type,
                "count": vent.count,
                "difficulty": vent.difficulty,
                "hours_per_unit": round(vent.hours_per_unit, 1),
                "total_hours": round(vent.total_hours, 1),
            })

    # ===================================================================
    # UNIT ITEMS (drains, penetration flashings, equipment)
    # Uses detailed vents if provided, else legacy counts
    # ===================================================================
    pipe_key, pipe_name = _PIPE_SEAL_KEY.get(system, ("Pipe_Boot_Seal", "Penetration Seal"))

    if m.vents:
        # Build unit items from detailed vent list
        _vent_to_pricing = {
            "pipe_boot": (pipe_key, pipe_name),
            "b_vent":    (pipe_key, f"B-Vent {pipe_name}"),
            "hood_vent": ("Gooseneck_Vent", "Vent Hood Flashing"),
            "plumb_vent": ("Plumbing_Vent", "Plumbing Vent Flashing"),
            "gum_box":   ("Gum_Box", "Gum Box / Catchment"),
            "scupper":   ("Scupper", "Overflow Scupper"),
            "radon_pipe": (pipe_key, f"Radon Pipe {pipe_name}"),
            "drain":     ("Roof_Drain", "Roof Drain Insert"),
        }
        for vent in m.vents:
            if vent.count <= 0:
                continue
            pkey_v, name_v = _vent_to_pricing.get(
                vent.vent_type, (pipe_key, vent.vent_type)
            )
            unit_price = _get_price(pkey_v)
            line_cost = vent.count * unit_price
            bid_grp = "roofing"

            results["unit_items"].append({
                "name": name_v,
                "base_count": vent.count,
                "multiplier": 1,
                "quantity": vent.count,
                "unit": "EA",
                "unit_price": round(unit_price, 2),
                "line_cost": round(line_cost, 2),
                "bid_group": bid_grp,
            })
            total_roofing_cost += line_cost
    else:
        # Legacy unit item definitions
        unit_defs = [
            ("Roof Drain Insert (OMG/Thaler)",
             "Roof_Drain", "EA", "roof_drain_count", 1, "roofing"),
            ("Overflow Scupper",
             "Scupper", "EA", "scupper_count", 1, "roofing"),
            ("Vent Hood Flashing",
             "Gooseneck_Vent", "EA", "vent_hood_count", 1, "roofing"),
            (f"Gas {pipe_name}",
             pipe_key, "EA", "gas_penetration_count", 1, "roofing"),
            (f"Electrical {pipe_name}",
             pipe_key, "EA", "electrical_penetration_count", 1, "roofing"),
            ("Plumbing Vent Flashing",
             "Plumbing_Vent", "EA", "plumbing_vent_count", 1, "roofing"),
            ("Gum Box / Catchment",
             "Gum_Box", "EA", "gum_box_count", 1, "roofing"),
            ("B-Vent Flashing",
             pipe_key, "EA", "b_vent_count", 1, "roofing"),
            ("Radon Pipe Seal",
             pipe_key, "EA", "radon_pipe_count", 1, "roofing"),
            ("Roof Hatch",
             "Roof_Hatch", "EA", "roof_hatch_count", 1, "roofing"),
        ]

        for name, pkey, unit, count_attr, multiplier, bid_grp in unit_defs:
            base_count = getattr(m, count_attr, 0)
            qty = base_count * multiplier
            if qty == 0:
                continue
            unit_price = _get_price(pkey)
            line_cost = qty * unit_price

            results["unit_items"].append({
                "name": name,
                "base_count": base_count,
                "multiplier": multiplier,
                "quantity": qty,
                "unit": unit,
                "unit_price": round(unit_price, 2),
                "line_cost": round(line_cost, 2),
                "bid_group": bid_grp,
            })

            if bid_grp == "roofing":
                total_roofing_cost += line_cost

    # Legacy curb flashing (when no detailed curbs)
    if not m.curbs:
        for name, pkey, unit, count_attr, mult, bid_grp in [
            ("Mechanical Unit Curb Flashing",
             "Flashing_General", "EA", "mechanical_unit_count", 4, "mechanical"),
            ("Sleeper Curb Flashing",
             "Flashing_General", "EA", "sleeper_curb_count", 2, "mechanical"),
        ]:
            base_count = getattr(m, count_attr, 0)
            qty = base_count * mult
            if qty == 0:
                continue
            unit_price = _get_price(pkey)
            line_cost = qty * unit_price
            results["unit_items"].append({
                "name": name,
                "base_count": base_count,
                "multiplier": mult,
                "quantity": qty,
                "unit": unit,
                "unit_price": round(unit_price, 2),
                "line_cost": round(line_cost, 2),
                "bid_group": bid_grp,
            })
            total_mechanical_cost += line_cost

    # Corner materials (Excel: corner count affects labour + material)
    if m.corner_count > 0:
        corner_price = _get_price("Flashing_General")
        corner_cost = m.corner_count * corner_price * 0.5  # half piece per corner
        results["unit_items"].append({
            "name": "Perimeter Corner Pieces",
            "base_count": m.corner_count,
            "multiplier": 1,
            "quantity": m.corner_count,
            "unit": "EA",
            "unit_price": round(corner_price * 0.5, 2),
            "line_cost": round(corner_cost, 2),
            "bid_group": "flashing",
        })
        total_flashing_cost += corner_cost

    # ===================================================================
    # CONSUMABLES — field-area based (Excel: FRS R41-R59)
    # ===================================================================
    consumable_defs = _SYSTEM_CONSUMABLES.get(system, _SYSTEM_CONSUMABLES["SBS"])

    for name, pkey, unit, rate_per_1000, bid_grp in consumable_defs:
        qty = math.ceil(roof_area / 1000 * rate_per_1000)
        if qty <= 0:
            qty = 1
        unit_price = _get_price(pkey)
        line_cost = qty * unit_price

        results["consumables"].append({
            "name": name,
            "quantity": qty,
            "unit": unit,
            "unit_price": round(unit_price, 2),
            "line_cost": round(line_cost, 2),
            "bid_group": bid_grp,
        })

        if bid_grp == "roofing":
            total_roofing_cost += line_cost
        elif bid_grp == "flashing":
            total_flashing_cost += line_cost

    # Wall-only consumables (adhesive/primer for parapet strips)
    wall_defs = _WALL_CONSUMABLES.get(system, [])
    wall_area = m.total_strip_sqft
    for name, pkey, unit, sqft_per_unit, bid_grp in wall_defs:
        if wall_area <= 0:
            continue
        qty = math.ceil(wall_area * 1.1 / sqft_per_unit)
        unit_price = _get_price(pkey)
        line_cost = qty * unit_price

        results["consumables"].append({
            "name": name,
            "quantity": qty,
            "unit": unit,
            "unit_price": round(unit_price, 2),
            "line_cost": round(line_cost, 2),
            "bid_group": bid_grp,
        })
        if bid_grp == "flashing":
            total_flashing_cost += line_cost

    # IKO Firetape / 6" Roof Tape — conditional on attachment method (Excel: FRS R53)
    if system == "SBS":
        any_torch_or_mop = m.vapour_barrier_attachment in ("Torched", "Mopped")
        if m.vapour_barrier_attachment == "Mopped":
            if m.vapour_barrier_product == "#15_Felt_x2":
                firetape_lf = parapet_lf
            else:
                # Densdeck quantity from area materials (if coverboard enabled)
                densdeck_qty = 0
                for am in results["area_materials"]:
                    if "Densdeck" in am.get("name", "") or "Soprasmart" in am.get("name", ""):
                        densdeck_qty = am.get("quantity", 0)
                        break
                firetape_lf = (densdeck_qty * 16) + 8 + parapet_lf
        else:
            firetape_lf = parapet_lf if any_torch_or_mop else 0

        if firetape_lf > 0:
            firetape_rolls = math.ceil(firetape_lf / 60)
            firetape_price = _get_price("Roof_Tape_IKO")
            firetape_cost = firetape_rolls * firetape_price
            results["consumables"].append({
                "name": "IKO Firetape 6\" (conditional on attachment)",
                "quantity": firetape_rolls,
                "unit": "roll (60 LF)",
                "unit_price": round(firetape_price, 2),
                "line_cost": round(firetape_cost, 2),
                "bid_group": "roofing",
            })
            total_roofing_cost += firetape_cost

    # PMMA System (Excel: FRS D31/D32) — Catalyst + Fleece
    if m.include_pmma and system == "SBS":
        # Sum perimeter section count for PMMA primer calc
        perim_section_count = sum(1 for s in m.perimeter_sections if s.lf > 0)
        # Catalyst: PMMA qty (from Alsan RS) × 7
        pmma_base_qty = math.ceil(roof_area / 100)  # approximate Alsan RS pail count
        catalyst_qty = pmma_base_qty * 7
        catalyst_price = _get_price("Catalyst")
        catalyst_cost = catalyst_qty * catalyst_price
        results["consumables"].append({
            "name": "PMMA Catalyst (Alsan RS)",
            "quantity": catalyst_qty,
            "unit": "can",
            "unit_price": round(catalyst_price, 2),
            "line_cost": round(catalyst_cost, 2),
            "bid_group": "roofing",
        })
        total_roofing_cost += catalyst_cost

        # Fleece: ROUNDUP(wall_area / 160, 0)
        pmma_wall_area = m.total_strip_sqft
        if pmma_wall_area > 0:
            fleece_qty = math.ceil(pmma_wall_area / 160)
            fleece_price = _get_price("Fleece_Reinforcement_Fabric")
            fleece_cost = fleece_qty * fleece_price
            results["consumables"].append({
                "name": "PMMA Fleece (Alsan RS)",
                "quantity": fleece_qty,
                "unit": "roll",
                "unit_price": round(fleece_price, 2),
                "line_cost": round(fleece_cost, 2),
                "bid_group": "roofing",
            })
            total_roofing_cost += fleece_cost

        # PMMA Primer: ROUNDUP(perim_section_count / 200, 0) pails
        if perim_section_count > 0:
            pmma_primer_qty = max(math.ceil(perim_section_count / 200), 1)
            pmma_primer_price = _get_price("Primer")
            pmma_primer_cost = pmma_primer_qty * pmma_primer_price
            results["consumables"].append({
                "name": "PMMA Primer (Alsan RS)",
                "quantity": pmma_primer_qty,
                "unit": "pail",
                "unit_price": round(pmma_primer_price, 2),
                "line_cost": round(pmma_primer_cost, 2),
                "bid_group": "roofing",
            })
            total_roofing_cost += pmma_primer_cost

    # Asphalt EasyMelt (Excel: FRS R46) — layer count from system
    if m.include_asphalt_easymelt and system == "SBS":
        layer_count = len(_SYSTEM_AREA_LAYERS.get(system, []))
        squares = roof_area / 100.0
        asphalt_qty = math.ceil(25 * squares / 50 * layer_count)
        asphalt_price = _get_price("Asphalt_EasyMelt")
        asphalt_cost = asphalt_qty * asphalt_price
        results["consumables"].append({
            "name": f"Asphalt EasyMelt ({layer_count} layers)",
            "quantity": asphalt_qty,
            "unit": "pail",
            "unit_price": round(asphalt_price, 2),
            "line_cost": round(asphalt_cost, 2),
            "bid_group": "roofing",
        })
        total_roofing_cost += asphalt_cost

    # Garland System (Excel: FRS R56-R59) — 4 products
    if m.garland_system:
        total_curbs = sum(c.count for c in m.curbs)
        # Tuff-Stuff MS: perimeter_lf / 15 tubes
        tuff_qty = math.ceil(parapet_lf / 15) if parapet_lf > 0 else 0
        if tuff_qty > 0:
            tuff_price = _get_price("Tuff_Stuff_MS")
            tuff_cost = tuff_qty * tuff_price
            results["consumables"].append({
                "name": "Tuff-Stuff MS (Garland)",
                "quantity": tuff_qty,
                "unit": "tube",
                "unit_price": round(tuff_price, 2),
                "line_cost": round(tuff_cost, 2),
                "bid_group": "flashing",
            })
            total_flashing_cost += tuff_cost

        # Gar-Mesh: ((perim/3)*3 + 4*curbs) / 150 * 1.1
        gar_mesh_area = ((parapet_lf / 3) * 3 + (4 * total_curbs)) / 150 * 1.1
        gar_mesh_rolls = max(math.ceil(gar_mesh_area), 1)
        gar_mesh_price = _get_price("Gar_Mesh")
        gar_mesh_cost = gar_mesh_rolls * gar_mesh_price
        results["consumables"].append({
            "name": "Gar-Mesh (Garland)",
            "quantity": gar_mesh_rolls,
            "unit": "roll",
            "unit_price": round(gar_mesh_price, 2),
            "line_cost": round(gar_mesh_cost, 2),
            "bid_group": "flashing",
        })
        total_flashing_cost += gar_mesh_cost

        # Garla-Flex: gar_mesh_rolls * 8/12 * 150 * 1.1 / 30
        garla_flex_pails = max(math.ceil(gar_mesh_rolls * 8 / 12 * 150 * 1.1 / 30), 1)
        garla_flex_price = _get_price("Garla_Flex")
        garla_flex_cost = garla_flex_pails * garla_flex_price
        results["consumables"].append({
            "name": "Garla-Flex (Garland)",
            "quantity": garla_flex_pails,
            "unit": "pail",
            "unit_price": round(garla_flex_price, 2),
            "line_cost": round(garla_flex_cost, 2),
            "bid_group": "flashing",
        })
        total_flashing_cost += garla_flex_cost

        # Flashing Bond Mastic: gar_mesh_rolls * 2
        mastic_pails = max(math.ceil(gar_mesh_rolls * 2), 1)
        mastic_price = _get_price("Flashing_Bond_Mastic_Garland")
        mastic_cost = mastic_pails * mastic_price
        results["consumables"].append({
            "name": "Flashing Bond Mastic (Garland)",
            "quantity": mastic_pails,
            "unit": "pail",
            "unit_price": round(mastic_price, 2),
            "line_cost": round(mastic_cost, 2),
            "bid_group": "flashing",
        })
        total_flashing_cost += mastic_cost

    # ===================================================================
    # EPDM / TPO SPECIFIC QUANTITY FORMULAS
    # (Excel: FRS R60-R101)
    # ===================================================================
    if system.startswith("EPDM"):
        # EPDM Seam Tape: membrane_rolls × seam overlap
        membrane_rolls = math.ceil(roof_area * 1.1 / 1000)
        seam_lf = roof_area / 10.0 * 1.1  # 10ft-wide rolls, seam every width
        seam_tape_rolls = math.ceil(seam_lf / 100.0)
        seam_tape_price = _get_price("EPDM_Seam_Tape")

        results["epdm_tpo_details"].append({
            "name": "EPDM Seam Tape (computed: roof_area/10 × 1.1 waste)",
            "quantity": seam_tape_rolls,
            "unit": "roll (100 lf)",
            "unit_price": round(seam_tape_price, 2),
            "line_cost": round(seam_tape_rolls * seam_tape_price, 2),
        })
        total_roofing_cost += seam_tape_rolls * seam_tape_price

        # EPDM Corners (inside + outside)
        total_corners = m.corner_count if m.corner_count > 0 else 4
        corner_price = _get_price("EPDM_PS_Corner")
        results["epdm_tpo_details"].append({
            "name": "EPDM Peel & Stick Corners (IS/OS)",
            "quantity": total_corners * 2,
            "unit": "piece",
            "unit_price": round(corner_price, 2),
            "line_cost": round(total_corners * 2 * corner_price, 2),
        })
        total_roofing_cost += total_corners * 2 * corner_price

        # EPDM Curb Flashing
        curb_perim = m.total_curb_perimeter_lf
        if curb_perim > 0:
            curb_flash_price = _get_price("EPDM_Curb_Flash")
            curb_flash_rolls = math.ceil(curb_perim / 50.0)  # 50 lf per roll
            results["epdm_tpo_details"].append({
                "name": "EPDM Curb Flash (from curb perimeters)",
                "quantity": curb_flash_rolls,
                "unit": "roll",
                "unit_price": round(curb_flash_price, 2),
                "line_cost": round(curb_flash_rolls * curb_flash_price, 2),
            })
            total_roofing_cost += curb_flash_rolls * curb_flash_price

        # EPDM RUSS-6 for perimeter
        russ_rolls = 0
        if parapet_lf > 0:
            russ_price = _get_price("EPDM_RUSS_6")
            russ_rolls = math.ceil(parapet_lf * 1.1 / 100.0)
            results["epdm_tpo_details"].append({
                "name": "EPDM RUSS 6\" (perimeter termination)",
                "quantity": russ_rolls,
                "unit": "roll",
                "unit_price": round(russ_price, 2),
                "line_cost": round(russ_rolls * russ_price, 2),
            })
            total_roofing_cost += russ_rolls * russ_price

        # HP-250 Primer — precise coverage (Excel: E79 formula)
        # = (seam_tape_rolls × 3/12 × 100) + (RUSS_rolls × 6/12 × 50 × 0.5) × 1.1
        hp250_area = (seam_tape_rolls * 3 / 12 * 100) + (russ_rolls * 6 / 12 * 50 * 0.5)
        hp250_area_with_waste = hp250_area * 1.1
        if hp250_area_with_waste > 0:
            hp250_gal = math.ceil(hp250_area_with_waste / 400)  # 400 sqft/gal
            hp250_price = _get_price("EPDM_Primer_HP250")
            hp250_cost = hp250_gal * hp250_price
            results["epdm_tpo_details"].append({
                "name": "EPDM Primer HP-250 (seam + RUSS area)",
                "quantity": hp250_gal,
                "unit": "gallon",
                "unit_price": round(hp250_price, 2),
                "line_cost": round(hp250_cost, 2),
            })
            total_roofing_cost += hp250_cost

    elif system.startswith("TPO"):
        # TPO 2nd membrane row (Excel: FRS R88)
        if m.tpo_second_membrane:
            tpo2_qty = math.ceil(roof_area * 1.1 / 1000)
            tpo2_price = _get_price("TPO_Membrane")
            tpo2_cost = tpo2_qty * tpo2_price
            results["epdm_tpo_details"].append({
                "name": "TPO Membrane 60 mil - 2nd Layer",
                "quantity": tpo2_qty,
                "unit": "roll (10'x100')",
                "unit_price": round(tpo2_price, 2),
                "line_cost": round(tpo2_cost, 2),
            })
            total_roofing_cost += tpo2_cost

        # TPO Rhinobond plate quantity (Excel: MAX(F25,F26,F28)×10)
        if system == "TPO_Mechanically_Attached":
            curb_perim = m.total_curb_perimeter_lf
            # Use max of active coverboard quantities (Securock/Densdeck/Fiberboard)
            coverboard_qtys = []
            for am in results["area_materials"]:
                nm = am.get("name", "")
                if any(k in nm for k in ("Securock", "Densdeck", "Soprasmart", "Fiberboard")):
                    coverboard_qtys.append(am.get("quantity", 0))
            max_cb = max(coverboard_qtys) if coverboard_qtys else math.ceil(roof_area * 1.1 / 32.0)
            rhinobond_qty = ((curb_perim + parapet_lf) + max_cb * 10) / 500.0 * 1.2
            rhinobond_pallets = math.ceil(rhinobond_qty) if rhinobond_qty > 0 else 1
            rb_price = _get_price("TPO_Rhinobond_Plate")
            results["epdm_tpo_details"].append({
                "name": "Rhinobond Plates (computed: edge + field)",
                "quantity": rhinobond_pallets,
                "unit": "pallet",
                "unit_price": round(rb_price, 2),
                "line_cost": round(rhinobond_pallets * rb_price, 2),
            })
            total_roofing_cost += rhinobond_pallets * rb_price

        # TPO Flashing — explicit toggles for 24" and 12" (Excel: FRS D90/D91)
        if parapet_lf > 0:
            for flash_size, include_flag, flash_key, flash_label in [
                ("24", m.include_tpo_flashing_24, "TPO_Flashing_24in", "TPO Flashing 24\" (parapet)"),
                ("12", m.include_tpo_flashing_12, "TPO_Flashing_12in", "TPO Flashing 12\" (parapet)"),
            ]:
                if include_flag:
                    flash_rolls = math.ceil(parapet_lf * 1.1 / 50)  # 50 lf per roll
                    flash_price = _get_price(flash_key)
                    results["epdm_tpo_details"].append({
                        "name": flash_label,
                        "quantity": flash_rolls,
                        "unit": "roll",
                        "unit_price": round(flash_price, 2),
                        "line_cost": round(flash_rolls * flash_price, 2),
                    })
                    total_flashing_cost += flash_rolls * flash_price

        # TPO Corners
        total_corners = m.corner_count if m.corner_count > 0 else 4
        tpo_corner_price = _get_price("TPO_Corner")
        results["epdm_tpo_details"].append({
            "name": "TPO Inside/Outside Corners",
            "quantity": total_corners * 2,
            "unit": "piece",
            "unit_price": round(tpo_corner_price, 2),
            "line_cost": round(total_corners * 2 * tpo_corner_price, 2),
        })
        total_roofing_cost += total_corners * 2 * tpo_corner_price

        # TPO Tuck Tape quantity (per seam LF)
        seam_lf = roof_area / 10.0 * 1.1
        tuck_rolls = math.ceil(seam_lf / 150.0)  # 150 lf per roll
        tuck_price = _get_price("TPO_Tuck_Tape")
        results["epdm_tpo_details"].append({
            "name": "TPO Tuck Tape (seam detail)",
            "quantity": tuck_rolls,
            "unit": "roll",
            "unit_price": round(tuck_price, 2),
            "line_cost": round(tuck_rolls * tuck_price, 2),
        })
        total_roofing_cost += tuck_rolls * tuck_price

    # ===================================================================
    # WOOD WORK (Excel: Takeoff R67-R76)
    # ===================================================================
    if m.wood_sections:
        for ws in m.wood_sections:
            qty = ws.quantity
            if qty <= 0:
                continue
            pkey = WOOD_PRODUCT_KEYS.get(ws.lumber_size, "Wood_Blocking_Lumber")
            unit_price = _get_price(pkey)
            line_cost = qty * unit_price
            unit_label = "4'x8' sheet" if ws.wood_type == "plywood" else "8ft piece"

            results["wood_materials"].append({
                "name": f"Wood: {ws.name} ({ws.wood_type}, {ws.lumber_size})",
                "quantity": qty,
                "unit": unit_label,
                "layers": ws.layers,
                "unit_price": round(unit_price, 2),
                "line_cost": round(line_cost, 2),
            })
            total_flashing_cost += line_cost

    # Wood facing from perimeter sections (parapet types with facing)
    if m.total_wood_face_sqft > 0:
        ply_sheets = math.ceil(m.total_wood_face_sqft / 32.0)
        ply_price = _get_price("Plywood_Sheathing")
        face_cost = ply_sheets * ply_price
        results["wood_materials"].append({
            "name": "Plywood Facing (parapet sections with facing)",
            "quantity": ply_sheets,
            "unit": "4'x8' sheet",
            "layers": 1,
            "unit_price": round(ply_price, 2),
            "line_cost": round(face_cost, 2),
        })
        total_flashing_cost += face_cost

    # ===================================================================
    # BATT INSULATION (Excel: Takeoff R77-R83)
    # ===================================================================
    if m.batt_sections:
        for bs in m.batt_sections:
            if bs.bundles <= 0:
                continue
            batt_price = _get_price("Batt_Insulation")
            line_cost = bs.bundles * batt_price
            results["batt_insulation"].append({
                "name": f"Batt Insulation: {bs.name} ({bs.insulation_type})",
                "sqft": round(bs.sqft, 0),
                "quantity": bs.bundles,
                "unit": "bundle",
                "layers": bs.layers,
                "unit_price": round(batt_price, 2),
                "line_cost": round(line_cost, 2),
            })
            total_roofing_cost += line_cost

    # ===================================================================
    # OTHER COSTS (Excel: FRS R123-R125)
    # Delivery, Disposal, Toilet, Fencing
    # ===================================================================
    # Delivery
    if m.delivery_count > 0:
        delivery_price = 250.00
        delivery_cost = m.delivery_count * delivery_price
        results["other_costs"].append({
            "name": "Delivery",
            "quantity": m.delivery_count,
            "unit": "trip",
            "unit_price": delivery_price,
            "line_cost": round(delivery_cost, 2),
        })
        total_other_cost += delivery_cost

    # Disposal
    if m.disposal_roof_count > 0:
        squares = roof_area / 100.0
        disposal_price = 70.00  # per square
        disposal_cost = m.disposal_roof_count * squares * disposal_price
        results["other_costs"].append({
            "name": f"Disposal ({m.disposal_roof_count} roof(s) x {squares:.0f} sq @ $70/sq)",
            "quantity": m.disposal_roof_count,
            "unit": "roof",
            "unit_price": round(squares * disposal_price, 2),
            "line_cost": round(disposal_cost, 2),
        })
        total_other_cost += disposal_cost

    # Toilet rental
    if m.include_toilet:
        toilet_cost = 250.00
        results["other_costs"].append({
            "name": "Portable Toilet Rental",
            "quantity": 1,
            "unit": "month",
            "unit_price": toilet_cost,
            "line_cost": toilet_cost,
        })
        total_other_cost += toilet_cost

    # Fencing
    if m.include_fencing:
        fencing_cost = 500.00 + (roof_area / 100.0 * 15.00)
        results["other_costs"].append({
            "name": "Temporary Fencing",
            "quantity": 1,
            "unit": "job",
            "unit_price": round(fencing_cost, 2),
            "line_cost": round(fencing_cost, 2),
        })
        total_other_cost += fencing_cost

    # ===================================================================
    # BID SUMMARY
    # ===================================================================
    total_roofing_plus_flashing = total_roofing_cost + total_flashing_cost
    labour_mult = meta["labour_multiplier"]
    mech_mult = meta["mechanical_multiplier"]

    results["bid_summary"] = {
        "item_1_general_requirements": {
            "description": "General Requirements (Div 01)",
            "note": "Mobilization, site protection, cleanup - typically 8-12% of roofing",
            "estimated_pct": 0.10,
            "estimated_cost": round(total_roofing_plus_flashing * 0.10, 2),
        },
        "item_2_roofing_assembly_and_flashing": {
            "description": f"Roofing Assembly + Metal Flashing ({meta['spec']})",
            "material_cost": round(total_roofing_plus_flashing, 2),
            "labour_multiplier": labour_mult,
            "note": meta["labour_note"],
            "estimated_cost": round(total_roofing_plus_flashing * labour_mult, 2),
        },
        "item_3_mechanical_support": {
            "description": "Mechanical Support (curbs, RTU flashings)",
            "material_cost": round(total_mechanical_cost, 2),
            "labour_multiplier": mech_mult,
            "note": "Higher labour ratio for detail work",
            "estimated_cost": round(total_mechanical_cost * mech_mult, 2),
        },
        "item_4_other_costs": {
            "description": "Other Costs (delivery, disposal, temp facilities)",
            "cost": round(total_other_cost, 2),
        },
        "total_material_cost": round(
            total_roofing_cost + total_flashing_cost + total_mechanical_cost, 2
        ),
        "total_other_cost": round(total_other_cost, 2),
        "total_estimate": round(
            total_roofing_plus_flashing * 0.10 +
            total_roofing_plus_flashing * labour_mult +
            total_mechanical_cost * mech_mult +
            total_other_cost,
            2
        ),
    }

    return results


# ---------------------------------------------------------------------------
# Reporting
# ---------------------------------------------------------------------------

def print_estimate(est: dict) -> None:
    """Pretty-print the quantity takeoff and cost estimate."""
    meas = est["project_measurements"]

    system_name = meas.get("roof_system_name", "Inverted Modified Bitumen (2-Ply SBS) - Soprema System")
    spec = meas.get("spec", "Div 07 52 01 / 07 62 00 / 07 92 00")
    print("=" * 72)
    print("  ROOFING QUANTITY TAKEOFF & COST ESTIMATE")
    print(f"  {system_name}")
    print(f"  Spec: {spec}  |  Layers: {meas.get('layer_count', '?')}")
    print("=" * 72)

    print(f"\n  Roof Area     : {meas['total_roof_area_sqft']:,.0f} sqft")
    print(f"  Perimeter     : {meas['perimeter_lf']:,.0f} LF")
    print(f"  Parapet       : {meas['parapet_length_lf']:,.0f} LF x {meas['parapet_height_ft']:.1f} ft")
    print(f"  Tapered Area  : {meas['tapered_area_sqft']:,.0f} sqft")
    print(f"  Ballast Area  : {meas['ballast_area_sqft']:,.0f} sqft")
    print(f"  Penetrations  : {meas['total_penetrations']} total")
    if meas.get("corner_count", 0) > 0:
        print(f"  Corners       : {meas['corner_count']}")

    # Roof sections
    if est.get("roof_sections"):
        print(f"\n  {'-' * 68}")
        print("  ROOF SECTIONS")
        print(f"  {'-' * 68}")
        for sec in est["roof_sections"]:
            print(f"    {sec['name']}: {sec['count']} x {sec['length_ft']}' x {sec['width_ft']}' = {sec['area_sqft']:,.0f} sqft")

    # Perimeter details
    if est.get("perimeter_details"):
        print(f"\n  {'-' * 68}")
        print("  PERIMETER SECTIONS (girth calculations)")
        print(f"  {'-' * 68}")
        for p in est["perimeter_details"]:
            print(f"    Section {p['name']}: {p['type']} | {p['height_in']}\"H x {p['lf']:,.0f} LF")
            print(f"      Strip: {p['strip_girth_in']}\" girth = {p['strip_sqft']:,.0f} sqft")
            print(f"      Metal: {p['metal_girth_in']}\" girth = {p['metal_sqft']:,.0f} sqft ({p['metal_sheets']} sheets)")
            if p.get('install_hours', 0) > 0:
                print(f"      Install: {p['install_hours']:.1f} hrs")

    # Curb details
    if est.get("curb_details"):
        print(f"\n  {'-' * 68}")
        print("  CURB DETAILS")
        print(f"  {'-' * 68}")
        for c in est["curb_details"]:
            print(f"    {c['curb_type']}: {c['count']}x {c['dimensions']}")
            print(f"      Perim: {c['perimeter_lf']:.1f} LF | Flash: {c['flashing_sqft']:.0f} sqft | Labour: {c['labour_hours']:.1f} hrs | Cost: ${c['flashing_cost']:,.2f}")

    # Vent details
    if est.get("vent_details"):
        print(f"\n  {'-' * 68}")
        print("  VENT DETAILS")
        print(f"  {'-' * 68}")
        for v in est["vent_details"]:
            print(f"    {v['vent_type']}: {v['count']}x ({v['difficulty']}) = {v['total_hours']:.1f} hrs")

    # Area materials
    print(f"\n  {'-' * 68}")
    print("  AREA-BASED MATERIALS (membrane, insulation, drainage)")
    print(f"  {'-' * 68}")
    for item in est["area_materials"]:
        cost_str = f"${item['line_cost']:,.2f}" if item["line_cost"] > 0 else "TBD"
        print(f"    {item['name']}")
        print(f"      {item['quantity']:,} {item['unit']}  @  ${item['unit_price']:,.2f}  =  {cost_str}")
        if item.get("note"):
            print(f"      ** {item['note']}")

    # Linear materials
    print(f"\n  {'-' * 68}")
    print("  LINEAR-FOOT MATERIALS (flashings, blocking, sheathing)")
    print(f"  {'-' * 68}")
    for item in est["linear_materials"]:
        print(f"    {item['name']}")
        print(f"      {item['quantity']:,} {item['unit']}  ({item['base_lf']:,.0f} LF + {item['waste_pct']} waste)")
        print(f"      @  ${item['unit_price']:,.2f}  =  ${item['line_cost']:,.2f}")

    # EPDM/TPO details
    if est.get("epdm_tpo_details"):
        print(f"\n  {'-' * 68}")
        print("  SYSTEM-SPECIFIC MATERIALS (EPDM/TPO)")
        print(f"  {'-' * 68}")
        for item in est["epdm_tpo_details"]:
            print(f"    {item['name']}")
            print(f"      {item['quantity']} {item['unit']}  @  ${item['unit_price']:,.2f}  =  ${item['line_cost']:,.2f}")

    # Unit items
    if est["unit_items"]:
        print(f"\n  {'-' * 68}")
        print("  UNIT ITEMS (drains, penetrations, equipment)")
        print(f"  {'-' * 68}")
        for item in est["unit_items"]:
            mult_str = f" x{item['multiplier']}" if item["multiplier"] > 1 else ""
            print(f"    {item['name']}")
            print(f"      {item['base_count']}{mult_str}  =  {item['quantity']} {item['unit']}")
            print(f"      @  ${item['unit_price']:,.2f}  =  ${item['line_cost']:,.2f}")

    # Consumables
    print(f"\n  {'-' * 68}")
    print("  CONSUMABLES & ACCESSORIES (field + wall)")
    print(f"  {'-' * 68}")
    for item in est["consumables"]:
        print(f"    {item['name']}")
        print(f"      {item['quantity']} {item['unit']}  @  ${item['unit_price']:,.2f}  =  ${item['line_cost']:,.2f}")

    # Wood materials
    if est.get("wood_materials"):
        print(f"\n  {'-' * 68}")
        print("  WOOD WORK MATERIALS")
        print(f"  {'-' * 68}")
        for item in est["wood_materials"]:
            print(f"    {item['name']}")
            print(f"      {item['quantity']} {item['unit']} x{item['layers']} layer(s)  @  ${item['unit_price']:,.2f}  =  ${item['line_cost']:,.2f}")

    # Batt insulation
    if est.get("batt_insulation"):
        print(f"\n  {'-' * 68}")
        print("  BATT INSULATION")
        print(f"  {'-' * 68}")
        for item in est["batt_insulation"]:
            print(f"    {item['name']}: {item['sqft']:,.0f} sqft = {item['quantity']} bundles  @  ${item['unit_price']:,.2f}  =  ${item['line_cost']:,.2f}")

    # Other costs
    if est.get("other_costs"):
        print(f"\n  {'-' * 68}")
        print("  OTHER COSTS")
        print(f"  {'-' * 68}")
        for item in est["other_costs"]:
            print(f"    {item['name']}: {item['quantity']} {item['unit']}  @  ${item['unit_price']:,.2f}  =  ${item['line_cost']:,.2f}")

    # Bid summary
    bid = est["bid_summary"]
    print(f"\n{'=' * 72}")
    print("  BID FORM SUMMARY")
    print(f"{'=' * 72}")

    item1 = bid["item_1_general_requirements"]
    print(f"\n  1. {item1['description']}")
    print(f"     ({item1['note']})")
    print(f"     Estimated: ${item1['estimated_cost']:>12,.2f}")

    item2 = bid["item_2_roofing_assembly_and_flashing"]
    print(f"\n  2. {item2['description']}")
    print(f"     Material: ${item2['material_cost']:>12,.2f}")
    print(f"     x {item2['labour_multiplier']}  ({item2['note']})")
    print(f"     Estimated: ${item2['estimated_cost']:>12,.2f}")

    item3 = bid["item_3_mechanical_support"]
    print(f"\n  3. {item3['description']}")
    print(f"     Material: ${item3['material_cost']:>12,.2f}")
    print(f"     x {item3['labour_multiplier']}  ({item3['note']})")
    print(f"     Estimated: ${item3['estimated_cost']:>12,.2f}")

    item4 = bid.get("item_4_other_costs", {})
    if item4.get("cost", 0) > 0:
        print(f"\n  4. {item4['description']}")
        print(f"     Cost:     ${item4['cost']:>12,.2f}")

    print(f"\n  {'-' * 50}")
    print(f"  TOTAL MATERIAL COST:     ${bid.get('total_material_cost', 0):>12,.2f}")
    if bid.get("total_other_cost", 0) > 0:
        print(f"  TOTAL OTHER COSTS:       ${bid['total_other_cost']:>12,.2f}")
    print(f"  TOTAL PROJECT ESTIMATE:  ${bid['total_estimate']:>12,.2f}")
    print(f"  {'-' * 50}")
    area = est['project_measurements']['total_roof_area_sqft']
    if area > 0:
        print(f"  Per sqft:  ${bid['total_estimate'] / area:>8,.2f} / sqft")
    print("=" * 72)


def export_json(est: dict, output_path: str) -> None:
    """Write the estimate to a JSON file."""
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(est, f, indent=2, ensure_ascii=False)
    print(f"\nJSON estimate saved to: {output_path}")


# ---------------------------------------------------------------------------
# AI Analysis Integration (drawing_analyzer.py output)
# ---------------------------------------------------------------------------

def load_analysis(json_path: str) -> dict:
    """Load AI analysis JSON produced by drawing_analyzer.py."""
    with open(json_path, "r", encoding="utf-8") as f:
        return json.load(f)


def measurements_from_analysis(analysis: dict,
                                total_roof_area_sqft: float,
                                perimeter_lf: float,
                                parapet_length_lf: float | None = None,
                                parapet_height_ft: float = 2.0) -> RoofMeasurements:
    """
    Build RoofMeasurements using AI-detected item counts + user-provided areas.

    The AI provides item counts from plan views; the user still provides
    area and perimeter (these must be measured from the scaled drawings).
    """
    counts: dict[str, int] = {}
    for plan in analysis.get("plan_analysis", []):
        if plan.get("parse_error"):
            continue
        for key, val in plan.get("counts", {}).items():
            counts[key] = counts.get(key, 0) + val

    return RoofMeasurements(
        total_roof_area_sqft=total_roof_area_sqft,
        perimeter_lf=perimeter_lf,
        parapet_length_lf=parapet_length_lf or perimeter_lf,
        parapet_height_ft=parapet_height_ft,
        roof_drain_count=counts.get("roof_drains", 0),
        scupper_count=counts.get("scuppers", 0),
        mechanical_unit_count=counts.get("mechanical_units", 0),
        sleeper_curb_count=counts.get("sleeper_curbs", 0),
        vent_hood_count=counts.get("vent_hoods", 0),
        gas_penetration_count=counts.get("gas_penetrations", 0),
        electrical_penetration_count=counts.get("electrical_penetrations", 0),
        plumbing_vent_count=counts.get("plumbing_vents", 0),
    )


def calculate_detail_takeoff(m: RoofMeasurements, analysis: dict) -> dict:
    """
    Calculate takeoff using AI-identified detail assemblies.

    Instead of hardcoded material lists, uses the detail_analysis from
    drawing_analyzer.py to determine which materials go in each detail,
    then applies measurements to calculate quantities and costs.

    Quantity resolution priority:
    1. Plan-view detail_quantities (AI measured from plan view)
    2. Detail-view scope_quantity (AI read from detail drawing annotations)
    3. DETAIL_TYPE_MAP fallback (global measurements, last resort)
    """
     #The calculate_detail_takeoff() function in backend/roof_estimator.py massively overestimates costs because it prices every AI-detected detail independently, applying the full roof area or perimeter to each. The reference Excel (231260__THE AMPERSAND 2026.xlsm) uses a single consolidated material list where each material appears once. This causes costs to be 3-5x what they should be.
     # line 2487 in roof_estimator
     # Example from the PDF output (26,500 sqft roof, 750 LF perimeter):
    results = {
        "project_measurements": {
            "total_roof_area_sqft": m.total_roof_area_sqft,
            "perimeter_lf": m.perimeter_lf,
            "parapet_length_lf": m.parapet_length_lf,
            "parapet_height_ft": m.parapet_height_ft,
            "total_penetrations": m.total_penetrations,
        },
        "details": [],
    }

    grand_total = 0.0

    # --- Build plan-view detail_quantities lookup ---
    # Merges detail_quantities from all plan pages into a single dict
    plan_detail_qtys: dict[str, dict] = {}
    for plan in analysis.get("plan_analysis", []):
        if plan.get("parse_error"):
            continue
        for ref_key, qty_info in plan.get("detail_quantities", {}).items():
            if isinstance(qty_info, dict) and qty_info.get("measurement", 0) > 0:
                plan_detail_qtys[ref_key] = qty_info

    # Collect all details from AI analysis
    all_details = []
    for page_data in analysis.get("detail_analysis", []):
        if page_data.get("parse_error"):
            continue
        drawing_ref = page_data.get("drawing_ref", "?")
        for detail in page_data.get("details", []):
            detail["_drawing_ref"] = drawing_ref
            all_details.append(detail)

    if not all_details:
        print("  WARNING: No AI detail analysis found. Use calculate_takeoff() instead.")
        return results

    # --- Detect alternative field assemblies (same drawing_ref = alternatives) ---
    field_assemblies = [d for d in all_details if d.get("detail_type") == "field_assembly"]
    if len(field_assemblies) > 1:
        # Group by drawing_ref — details on the same page are alternatives
        primary_ref = field_assemblies[0].get("_drawing_ref")
        for fa in field_assemblies[1:]:
            if fa.get("_drawing_ref") == primary_ref:
                fa["_is_alternative"] = True

    for detail in all_details:
        dtype = detail.get("detail_type", "unknown")
        dname = detail.get("detail_name", "Unknown Detail")
        dref = detail.get("_drawing_ref", "?")
        quantity_source = "fallback"

        # --- Resolve base_value and mtype using priority order ---

        # Priority 1: Plan-view detail_quantities
        # Try matching by detail name pattern (e.g., "Detail 5/R3.1")
        plan_qty = None
        detail_num = dname.split(" - ")[0].strip() if " - " in dname else dname
        for ref_key, qty_info in plan_detail_qtys.items():
            # Match "Detail 5/R3.1" against plan keys like "Detail 5/R3.1"
            if detail_num.lower() in ref_key.lower() or ref_key.lower() in f"{detail_num}/{dref}".lower():
                plan_qty = qty_info
                break

        if plan_qty and plan_qty.get("measurement", 0) > 0:
            base_value = float(plan_qty["measurement"])
            mtype = plan_qty.get("unit", detail.get("measurement_type", "each"))
            quantity_source = "plan_view"

        # Priority 2: AI scope_quantity from detail drawing
        elif detail.get("scope_quantity") is not None and detail["scope_quantity"] > 0:
            base_value = float(detail["scope_quantity"])
            mtype = detail.get("scope_unit", detail.get("measurement_type", "each"))
            quantity_source = "detail_drawing"

        # Priority 3: DETAIL_TYPE_MAP fallback (global measurements)
        else:
            mtype = detail.get("measurement_type", "each")
            type_info = DETAIL_TYPE_MAP.get(dtype)
            if type_info:
                map_mtype, attr = type_info
                base_value = getattr(m, attr, 0)
                # Use the map's measurement_type to stay consistent with base_value
                mtype = map_mtype
            else:
                base_value = 1

            # --- Adjustments for fallback mode ---
            if dtype == "expansion_joint":
                base_value = base_value * 0.25

            # Use realistic curb perimeters instead of arbitrary multipliers
            if dtype in CURB_TYPICAL_PERIMETER_LF and mtype == "linear_ft":
                base_value = base_value * CURB_TYPICAL_PERIMETER_LF[dtype]

        detail_result = {
            "detail_name": dname,
            "detail_type": dtype,
            "drawing_ref": dref,
            "measurement_type": mtype,
            "base_measurement": round(base_value, 1),
            "quantity_source": quantity_source,
            "layers": [],
            "detail_cost": 0.0,
        }

        # --- Fix 1: Skip alternative field assemblies (keep for reference only) ---
        if detail.get("_is_alternative"):
            detail_result["is_alternative"] = True
            detail_result["detail_cost"] = 0.0
            detail_result["note"] = "Alternative assembly — not included in total"
            results["details"].append(detail_result)
            continue

        # --- Fix 3: Deduplicate repeated material layers by pricing_key ---
        seen_pkeys: set[str] = set()
        deduped_layers: list[dict] = []
        for layer in detail.get("layers", []):
            pkey = layer.get("pricing_key", "CUSTOM")
            if pkey != "CUSTOM" and pkey in seen_pkeys:
                continue  # skip duplicate material
            seen_pkeys.add(pkey)
            deduped_layers.append(layer)

        for layer in deduped_layers:
            pkey = layer.get("pricing_key", "CUSTOM")
            material_name = layer.get("material", "?")
            notes = layer.get("notes", "")

            # Safely parse the AI-extracted cross-sectional dimension
            raw_dim = layer.get("dimension_in")
            try:
                dimension_in = float(raw_dim) if raw_dim is not None else 0.0
            except (ValueError, TypeError):
                dimension_in = 0.0

            # Price resolution logic
            price = _get_price(pkey)
            if pkey == "CUSTOM" or price == 0.0:
                detail_result["layers"].append({
                    "material": material_name,
                    "pricing_key": pkey,
                    "quantity": "?",
                    "unit": "?",
                    "unit_price": 0.0,
                    "line_cost": 0.0,
                    "notes": notes,
                    "warning": f"No pricing for '{pkey}' - needs manual entry",
                })
                continue

            unit_price = price
            coverage = COVERAGE_RATES.get(pkey, {})
            waste = 1.10  # 10% standard waste factor
            mat_scope = _material_scope(pkey)

            # ---------------------------------------------------------------
            # Guard 1: Intrinsic unit items (per_each) must always use the
            # "each" path, regardless of the parent detail's measurement type.
            # Without this, a Roof_Drain nested inside a field_assembly detail
            # (mtype="sqft", base=6777) would compute qty = ceil(6777/32) = 233.
            # ---------------------------------------------------------------
            if mat_scope == "discrete":
                if mtype == "each":
                    qty = max(1, int(base_value * coverage.get("per_each", 1)))
                else:
                    qty = 1
                unit = coverage.get("unit", "EA")

            # ---------------------------------------------------------------
            # Guard 2: Completely unknown key (not in COVERAGE_RATES) embedded
            # in an area/linear detail — avoid the default /32 sqft fallback
            # which causes massive over-counts for things like HVAC_Curb_Detail.
            # ---------------------------------------------------------------
            elif not coverage and mtype in ("sqft", "linear_ft"):
                qty = 1
                unit = "EA"

            # ---------------------------------------------------------------
            # Fix 2: For sqft details, linear materials use perimeter not area
            # ---------------------------------------------------------------
            elif mtype == "sqft" and mat_scope == "linear":
                # Linear materials (flashing, sealant, etc.) inside a sqft detail
                # only apply at edges/perimeter, NOT across the full area.
                perimeter = m.perimeter_lf or (4 * math.sqrt(base_value))
                lf_per = coverage.get("lf_per_unit", 1)
                qty = math.ceil(perimeter * waste / lf_per)
                unit = coverage.get("unit", "unit")

            elif mtype == "sqft":
                # Area materials: apply to full base area (correct behavior)
                sqft_per = coverage.get("sqft_per_unit", 32)
                qty = math.ceil(base_value * waste / sqft_per)
                unit = coverage.get("unit", "unit")

            # True-area calculation for linear details
            elif mtype == "linear_ft":
                if dimension_in > 0:
                    # Transform 1D linear run into 2D area required for the specific material layer
                    girth_ft = dimension_in / 12.0
                    actual_sqft = base_value * girth_ft

                    if "sqft_per_unit" in coverage:
                        sqft_per = coverage.get("sqft_per_unit", 32)
                        qty = math.ceil(actual_sqft * waste / sqft_per)
                        unit = coverage.get("unit", "unit")
                    elif "lf_per_unit" in coverage:
                        # Handle strictly linear products (e.g., metal cap flashing)
                        lf_per = coverage.get("lf_per_unit", 1)
                        qty = math.ceil(base_value * waste / lf_per)
                        unit = coverage.get("unit", "unit")
                    else:
                        qty = max(1, int(base_value * waste))
                        unit = coverage.get("unit", "EA")
                else:
                    # Fallback heuristic if the vision model failed to extract a dimension
                    if "sqft_per_unit" in coverage and "lf_per_unit" not in coverage:
                        sqft_per = coverage.get("sqft_per_unit", 32)
                        qty = math.ceil(base_value * waste / sqft_per)
                        unit = coverage.get("unit", "unit")
                    elif "lf_per_unit" in coverage:
                        lf_per = coverage.get("lf_per_unit", 1)
                        qty = math.ceil(base_value * waste / lf_per)
                        unit = coverage.get("unit", "unit")
                    else:
                        qty = max(1, int(base_value))
                        unit = coverage.get("unit", "EA")

            else:  # Discrete item logic (each)
                per_each = coverage.get("per_each", 1)
                qty = max(1, int(base_value * per_each))
                unit = coverage.get("unit", "EA")

            # ---------------------------------------------------------------
            # Fix 5: Sanity cap — for "each" type details, area/linear
            # materials shouldn't produce enormous quantities (a single
            # penetration or curb has a small footprint, not the whole roof).
            # ---------------------------------------------------------------
            if mtype == "each" and mat_scope == "area" and base_value <= 20:
                # Cap: assume max ~100 sqft of material per unit for discrete details
                sqft_per = coverage.get("sqft_per_unit", 32)
                max_qty = math.ceil(100 * base_value * waste / sqft_per)
                qty = min(qty, max(1, max_qty))

            line_cost = qty * unit_price

            detail_result["layers"].append({
                "material": material_name,
                "pricing_key": pkey,
                "quantity": qty,
                "unit": unit,
                "unit_price": round(unit_price, 2),
                "line_cost": round(line_cost, 2),
                "notes": notes,
            })
            detail_result["detail_cost"] += line_cost

        detail_result["detail_cost"] = round(detail_result["detail_cost"], 2)

        # --- Sanity cap: flag details that exceed $100/sqft of roof area ---
        roof_area = m.total_roof_area_sqft or 1
        max_reasonable_detail_cost = roof_area * 100  # $100/sqft is extremely high
        if detail_result["detail_cost"] > max_reasonable_detail_cost:
            detail_result["warning"] = (
                f"Detail cost ${detail_result['detail_cost']:,.0f} exceeds sanity cap "
                f"(${max_reasonable_detail_cost:,.0f}). Quantity source: {quantity_source}. "
                f"Capping at ${max_reasonable_detail_cost:,.0f}."
            )
            detail_result["detail_cost_uncapped"] = detail_result["detail_cost"]
            detail_result["detail_cost"] = round(max_reasonable_detail_cost, 2)

        grand_total += detail_result["detail_cost"]
        results["details"].append(detail_result)

    results["total_material_cost"] = round(grand_total, 2)
    results["bid_summary"] = {
        "material_cost": round(grand_total, 2),
    }

    return results


# ---------------------------------------------------------------------------
# Integration Engine — join_takeoff_data()
# ---------------------------------------------------------------------------
# Merges two independent data streams:
#   spatial_json  — output of drawing_analyzer.py  (Vision LLM, spatial data)
#   spec_json     — output of file_extractor.py    (deterministic regex)
#
# The join is strictly deterministic:
#   • Quantities come exclusively from spatial_json (LLM extracted dims/counts)
#   • Material identities come exclusively from spec_json["spec_materials"]
#   • Any spatial detail that cannot be matched to a confirmed spec material
#     is flagged as a "Material Resolution Failure" at WARNING level.
# ---------------------------------------------------------------------------

# Maps drawing detail_type → the spec_json pricing_key(s) that satisfy it.
# A detail_type may accept multiple pricing_keys (ordered by preference).
# The first key found in spec_materials wins.
_DETAIL_TYPE_TO_SPEC_KEYS: dict[str, list[str]] = {
    "field_assembly": [
        "Polyisocyanurate_ISO_Insulation",
        "Tapered_ISO",
        "TPO_Membrane",
        "EPDM_Membrane",
        "SBS_Membrane",
        "Cap_Membrane",
        "Base_Membrane",
        "TPO_60mil_Mechanically_Attached",
        "SBS_2Ply_Modified_Bitumen",
        "EPDM_60mil_Fully_Adhered",
    ],
    "parapet": [
        "Flashing_General",
        "Metal_Flashing_Galvanized",
        "Metal_Flashing_Prepainted",
        "Cap_Membrane",
        "Base_Membrane",
    ],
    "curtain_wall": [
        "Flashing_General",
        "Metal_Flashing_Galvanized",
        "Metal_Flashing_Prepainted",
    ],
    "drain": [
        "Roof_Drain",
    ],
    "mechanical_curb": [
        "Flashing_General",
        "Metal_Flashing_Galvanized",
    ],
    "sleeper_curb": [
        "Wood_Blocking_Lumber",
        "Plywood_Sheathing",
    ],
    "penetration_gas": [
        "Pipe_Boot_Seal",
        "EPDM_Pipe_Flashing",
        "TPO_Pipe_Boot",
        "Gooseneck_Vent",
    ],
    "penetration_electrical": [
        "Pipe_Boot_Seal",
        "EPDM_Pipe_Flashing",
        "TPO_Pipe_Boot",
    ],
    "penetration_plumbing": [
        "Plumbing_Vent",
        "Pipe_Boot_Seal",
        "EPDM_Pipe_Flashing",
        "TPO_Pipe_Boot",
    ],
    "vent_hood": [
        "Gooseneck_Vent",
        "Vent_Cap",
    ],
    "scupper": [
        "Scupper",
    ],
    "expansion_joint": [
        "Flashing_General",
        "EPDM_Accessory",
        "TPO_Accessory",
    ],
    "pipe_support": [
        "Wood_Blocking_Lumber",
        "Clips",
    ],
    "opening_cover": [
        "Plywood_Sheathing",
        "Flashing_General",
    ],
}


def join_takeoff_data(
    spatial_json: dict,
    spec_json: dict,
) -> dict:
    """
    Deterministically merge spatial drawing data with spec material data.

    Parameters
    ----------
    spatial_json : dict
        Output of ``drawing_analyzer.py::analyze_drawing()``.
        Must contain ``plan_analysis`` and/or ``detail_analysis`` keys.

    spec_json : dict
        Output of ``file_extractor.py::analyze_text()``.
        Must contain a ``spec_materials`` key — a flat dict keyed by
        canonical PRICING keys (e.g. ``"Tapered_ISO"``, ``"Roof_Drain"``).

    Returns
    -------
    dict with the following structure::

        {
          "resolved_line_items": [
            {
              "detail_name": str,
              "detail_type": str,
              "pricing_key": str,
              "material_name": str,
              "quantity": float,
              "unit": str,
              "unit_price": float,
              "line_cost": float,
              "quantity_source": str,  # "plan_view" | "detail_drawing" | "fallback"
              "spec_pages": [int, ...]  # pages in spec doc where material confirmed
            }
          ],
          "material_failures": [
            {
              "detail_name": str,
              "detail_type": str,
              "expected_pricing_keys": [str, ...],
              "message": str
            }
          ],
          "bid_summary": {
            "total_material_cost": float,
            "total_line_items": int,
            "total_failures": int
          }
        }

    Raises
    ------
    ValueError
        If ``spec_json`` is missing the ``spec_materials`` key (wrong format).
    """
    if "spec_materials" not in spec_json:
        raise ValueError(
            "spec_json is missing the 'spec_materials' key. "
            "Ensure it was produced by file_extractor.py::analyze_text()."
        )

    confirmed_spec: dict[str, dict] = spec_json["spec_materials"]  # pricing_key → info
    resolved_line_items: list[dict] = []
    material_failures: list[dict] = []
    grand_total: float = 0.0

    # ------------------------------------------------------------------
    # Build plan-view detail_quantities lookup (same logic as calculate_detail_takeoff)
    # ------------------------------------------------------------------
    plan_detail_qtys: dict[str, dict] = {}
    for plan in spatial_json.get("plan_analysis", []):
        if plan.get("parse_error"):
            continue
        for ref_key, qty_info in plan.get("detail_quantities", {}).items():
            if isinstance(qty_info, dict) and qty_info.get("measurement", 0) > 0:
                plan_detail_qtys[ref_key] = qty_info

    # Aggregate simple item counts from all plan pages
    item_counts: dict[str, int] = {}
    for plan in spatial_json.get("plan_analysis", []):
        if plan.get("parse_error"):
            continue
        for k, v in plan.get("counts", {}).items():
            item_counts[k] = item_counts.get(k, 0) + v

    # ------------------------------------------------------------------
    # Iterate over all AI-identified details
    # ------------------------------------------------------------------
    all_details: list[dict] = []
    for page_data in spatial_json.get("detail_analysis", []):
        if page_data.get("parse_error"):
            continue
        drawing_ref = page_data.get("drawing_ref", "?")
        for detail in page_data.get("details", []):
            detail["_drawing_ref"] = drawing_ref
            all_details.append(detail)

    if not all_details:
        logger.warning(
            "[join_takeoff_data] No detail_analysis entries in spatial_json. "
            "Proceeding with plan-view counts only."
        )

    field_assembly_count = sum(
        1 for d in all_details if d.get("detail_type") == "field_assembly"
    )

    for detail in all_details:
        dtype: str = detail.get("detail_type", "unknown")
        dname: str = detail.get("detail_name", "Unknown Detail")
        dref: str = detail.get("_drawing_ref", "?")

        # ------------------------------------------------------------------
        # Resolve quantity (mirrors priority order in calculate_detail_takeoff)
        # ------------------------------------------------------------------
        quantity_source: str = "fallback"
        base_value: float = 1.0
        mtype: str = detail.get("measurement_type", "each")

        # Priority 1: plan_detail_quantities
        detail_num = dname.split(" - ")[0].strip() if " - " in dname else dname
        plan_qty: dict | None = None
        for ref_key, qty_info in plan_detail_qtys.items():
            if (detail_num.lower() in ref_key.lower()
                    or ref_key.lower() in f"{detail_num}/{dref}".lower()):
                plan_qty = qty_info
                break

        if plan_qty and plan_qty.get("measurement", 0) > 0:
            base_value = float(plan_qty["measurement"])
            mtype = plan_qty.get("unit", mtype)
            quantity_source = "plan_view"

        # Priority 2: AI scope_quantity from detail drawing
        elif (detail.get("scope_quantity") is not None
              and detail["scope_quantity"] > 0):
            base_value = float(detail["scope_quantity"])
            mtype = detail.get("scope_unit", mtype)
            quantity_source = "detail_drawing"

        # Priority 3: DETAIL_TYPE_MAP fallback
        else:
            type_info = DETAIL_TYPE_MAP.get(dtype)
            if type_info:
                map_mtype, _ = type_info
                mtype = map_mtype
            base_value = 1.0  # conservative — no spatial data available
            quantity_source = "fallback"

        if dtype == "field_assembly" and field_assembly_count > 1:
            base_value = base_value / field_assembly_count

        # ------------------------------------------------------------------
        # Resolve material from spec_json  (DETERMINISTIC — no LLM)
        # ------------------------------------------------------------------
        candidate_keys: list[str] = _DETAIL_TYPE_TO_SPEC_KEYS.get(dtype, [])

        # Also include any pricing_keys suggested by the AI detail layers
        for layer in detail.get("layers", []):
            pk = layer.get("pricing_key", "")
            if pk and pk not in candidate_keys:
                candidate_keys = [pk] + candidate_keys  # AI suggestion takes priority

        matched_key: str | None = None
        for candidate in candidate_keys:
            if candidate in confirmed_spec:
                matched_key = candidate
                break

        if matched_key is None:
            # ---- Material Resolution Failure ----
            failure_msg = (
                f"Material Resolution Failure: Detail '{dname}' (type={dtype}, "
                f"ref={dref}) requires one of {candidate_keys} but none were "
                f"confirmed in the Specification document (spec_materials is empty "
                f"or does not contain a matching key). "
                f"Verify Specification PDF coverage or add missing regex patterns "
                f"to database.PRODUCT_KEYWORDS."
            )
            logger.warning(failure_msg)
            material_failures.append({
                "detail_name": dname,
                "detail_type": dtype,
                "drawing_ref": dref,
                "expected_pricing_keys": list(candidate_keys),
                "message": failure_msg,
            })
            continue  # skip — cannot price without confirmed material

        # ------------------------------------------------------------------
        # Calculate quantity and cost (deterministic)
        # ------------------------------------------------------------------
        spec_info: dict = confirmed_spec[matched_key]
        unit_price: float = _get_price(matched_key)
        coverage: dict = COVERAGE_RATES.get(matched_key, {})
        waste: float = 1.10

        if mtype == "sqft":
            sqft_per = float(coverage.get("sqft_per_unit", 32))
            quantity = math.ceil(base_value * waste / sqft_per)
            unit = coverage.get("unit", "unit")
        elif mtype == "linear_ft":
            if "lf_per_unit" in coverage:
                lf_per = float(coverage["lf_per_unit"])
                quantity = math.ceil(base_value * waste / lf_per)
                unit = coverage.get("unit", "unit")
            else:
                # Treat as sqft fallback
                sqft_per = float(coverage.get("sqft_per_unit", 32))
                quantity = math.ceil(base_value * waste / sqft_per)
                unit = coverage.get("unit", "unit")
        else:  # each
            per_each = coverage.get("per_each", 1)
            quantity = max(1, int(base_value * per_each))
            unit = coverage.get("unit", "EA")

        line_cost: float = quantity * unit_price
        grand_total += line_cost

        resolved_line_items.append({
            "detail_name": dname,
            "detail_type": dtype,
            "drawing_ref": dref,
            "pricing_key": matched_key,
            "material_name": spec_info["product_name"],
            "quantity": quantity,
            "unit": unit,
            "unit_price": round(unit_price, 2),
            "line_cost": round(line_cost, 2),
            "quantity_source": quantity_source,
            "spec_pages": spec_info.get("pages", []),
        })

    # ------------------------------------------------------------------
    # Log summary
    # ------------------------------------------------------------------
    logger.info(
        "[join_takeoff_data] Resolved %d line items | %d Material Resolution Failures | "
        "Total material cost: $%,.2f",
        len(resolved_line_items),
        len(material_failures),
        grand_total,
    )
    if material_failures:
        logger.warning(
            "[join_takeoff_data] %d detail(s) could not be resolved to a confirmed "
            "spec material. Review 'material_failures' in the return value.",
            len(material_failures),
        )

    return {
        "resolved_line_items": resolved_line_items,
        "material_failures": material_failures,
        "bid_summary": {
            "total_material_cost": round(grand_total, 2),
            "total_line_items": len(resolved_line_items),
            "total_failures": len(material_failures),
        },
    }


def print_join_result(join_result: dict) -> None:
    """Pretty-print the output of join_takeoff_data()."""
    bid = join_result["bid_summary"]
    items = join_result["resolved_line_items"]
    failures = join_result["material_failures"]

    print("\n" + "=" * 72)
    print("  INTEGRATED TAKEOFF  (Spatial + Spec — Deterministic Join)")
    print("=" * 72)

    if items:
        print(f"\n  {'Detail':<40} {'Material':<30} {'Qty':>6} {'Unit':<14} {'Cost':>12}")
        print(f"  {'-' * 68}")
        for item in items:
            src_tag = f"[{item['quantity_source'][:3].upper()}]"
            print(
                f"  {item['detail_name'][:39]:<40} "
                f"{item['material_name'][:29]:<30} "
                f"{item['quantity']:>6} "
                f"{item['unit'][:13]:<14} "
                f"${item['line_cost']:>10,.2f}  {src_tag}"
            )

    if failures:
        print(f"\n  {'─' * 68}")
        print(f"  MATERIAL RESOLUTION FAILURES  ({len(failures)} total)")
        print(f"  {'─' * 68}")
        for f in failures:
            print(f"  !! {f['detail_name']}  [{f['detail_type']}]")
            print(f"     Expected one of: {f['expected_pricing_keys']}")
            print(f"     → Not confirmed in Specification document.")

    print(f"\n  {'─' * 68}")
    print(f"  Resolved line items : {bid['total_line_items']}")
    print(f"  Resolution failures : {bid['total_failures']}")
    print(f"  Total material cost : ${bid['total_material_cost']:>12,.2f}")
    print("=" * 72)


def print_detail_estimate(est: dict) -> None:
    """Pretty-print a detail-based estimate (from AI analysis)."""
    meas = est["project_measurements"]

    print("=" * 72)
    print("  DETAIL-BASED QUANTITY TAKEOFF (AI-Analyzed)")
    print("=" * 72)

    print(f"\n  Roof Area     : {meas['total_roof_area_sqft']:,.0f} sqft")
    print(f"  Perimeter     : {meas['perimeter_lf']:,.0f} LF")
    print(f"  Parapet       : {meas['parapet_length_lf']:,.0f} LF x {meas['parapet_height_ft']:.1f} ft")
    print(f"  Penetrations  : {meas['total_penetrations']} total")

    for detail in est.get("details", []):
        dtype = detail["detail_type"]
        mtype = detail["measurement_type"]
        base = detail["base_measurement"]
        cost = detail["detail_cost"]

        print(f"\n  {'-' * 68}")
        print(f"  {detail['detail_name']}  [{dtype}]  (ref: {detail.get('drawing_ref', '?')})")
        print(f"  Measured in: {mtype}  |  Base value: {base:,}  |  Detail cost: ${cost:,.2f}")
        print(f"  {'-' * 68}")

        for layer in detail.get("layers", []):
            warning = f"  !! {layer['warning']}" if layer.get("warning") else ""
            if layer.get("line_cost", 0) > 0:
                print(f"    {layer['material']}")
                print(f"      {layer['quantity']:,} {layer['unit']}  @  ${layer['unit_price']:,.2f}  =  ${layer['line_cost']:,.2f}")
                if layer.get("notes"):
                    print(f"      ({layer['notes']})")
            else:
                print(f"    {layer['material']}  ->  {layer['pricing_key']}{warning}")
                if layer.get("notes"):
                    print(f"      ({layer['notes']})")

    bid = est.get("bid_summary", {})
    if bid:
        print(f"\n{'=' * 72}")
        print("  ESTIMATE SUMMARY")
        print(f"{'=' * 72}")
        print(f"  Total Material Cost:       ${bid.get('material_cost', 0):>12,.2f}")
        print(f"  Labour Cost:               ${bid.get('labour_cost', 0):>12,.2f}")
        print(f"  Other Costs:               ${bid.get('other_costs', 0):>12,.2f}")
        print(f"  {'-' * 50}")
        print(f"  Total Direct Cost (COGS):  ${bid.get('total_direct_cost', 0):>12,.2f}")
        print(f"  Overhead (35%):            ${bid.get('overhead_35pct', 0):>12,.2f}")
        print(f"  Breakeven:                 ${bid.get('breakeven', 0):>12,.2f}")
        print(f"  Net Profit (10%):          ${bid.get('profit_10pct', 0):>12,.2f}")
        print(f"  {'=' * 50}")
        print(f"  SELLING PRICE:             ${bid.get('total_estimate', 0):>12,.2f}")
        print(f"  Per sqft:                  ${bid.get('per_sqft', 0):>12,.2f}")
    print("=" * 72)


# ---------------------------------------------------------------------------
# CLI - interactive measurement input
# ---------------------------------------------------------------------------

def _input_float(prompt: str, default: float | None = None) -> float:
    """Prompt for a float with optional default."""
    suffix = f" [{default}]" if default is not None else ""
    while True:
        raw = input(f"  {prompt}{suffix}: ").strip()
        if not raw and default is not None:
            return default
        try:
            return float(raw)
        except ValueError:
            print("    Please enter a number.")


def _input_int(prompt: str, default: int = 0) -> int:
    """Prompt for an integer with default."""
    while True:
        raw = input(f"  {prompt} [{default}]: ").strip()
        if not raw:
            return default
        try:
            return int(raw)
        except ValueError:
            print("    Please enter a whole number.")


def main():
    json_output = None
    analysis_path = None

    if "--json" in sys.argv:
        idx = sys.argv.index("--json")
        if idx + 1 < len(sys.argv):
            json_output = sys.argv[idx + 1]

    if "--analysis" in sys.argv:
        idx = sys.argv.index("--analysis")
        if idx + 1 < len(sys.argv):
            analysis_path = sys.argv[idx + 1]

    print("=" * 60)
    print("  ROOF ESTIMATOR - Drawing-Based Quantity Takeoff")
    print("=" * 60)

    # Always need area + perimeter from the user (scaled drawings)
    print("\nEnter measurements from scaled architectural drawings.\n")

    print("[AREAS]")
    area = _input_float("Total roof area (sqft)")
    perim = _input_float("Roof perimeter (LF)")

    print("\n[PARAPET]")
    par_len = _input_float("Parapet length (LF)", default=perim)
    par_ht = _input_float("Parapet height (ft)", default=2.0)

    if analysis_path:
        # --- AI-DRIVEN MODE: use Gemini analysis for details + counts ---
        print(f"\nLoading AI analysis from: {analysis_path}")
        analysis = load_analysis(analysis_path)

        # Get counts from AI, but still ask for area/perimeter
        measurements = measurements_from_analysis(
            analysis, area, perim, par_len, par_ht
        )

        print(f"\nAI-detected counts:")
        print(f"  Drains: {measurements.roof_drain_count}  |  Scuppers: {measurements.scupper_count}")
        print(f"  Mechanical: {measurements.mechanical_unit_count}  |  Sleepers: {measurements.sleeper_curb_count}")
        print(f"  Vents: {measurements.vent_hood_count}  |  Gas: {measurements.gas_penetration_count}")
        print(f"  Electrical: {measurements.electrical_penetration_count}  |  Plumbing: {measurements.plumbing_vent_count}")

        override = input("\n  Override any counts? (y/N): ").strip().lower()
        if override == "y":
            measurements.roof_drain_count = _input_int("Roof drains", measurements.roof_drain_count)
            measurements.scupper_count = _input_int("Scuppers", measurements.scupper_count)
            measurements.mechanical_unit_count = _input_int("Mechanical units", measurements.mechanical_unit_count)
            measurements.sleeper_curb_count = _input_int("Sleeper curbs", measurements.sleeper_curb_count)
            measurements.vent_hood_count = _input_int("Vent hoods", measurements.vent_hood_count)
            measurements.gas_penetration_count = _input_int("Gas penetrations", measurements.gas_penetration_count)
            measurements.electrical_penetration_count = _input_int("Electrical penetrations", measurements.electrical_penetration_count)
            measurements.plumbing_vent_count = _input_int("Plumbing vents", measurements.plumbing_vent_count)

        print(f"\nCalculating detail-based estimate...\n")
        estimate = calculate_detail_takeoff(measurements, analysis)
        print_detail_estimate(estimate)

        # Also run the standard estimate for comparison
        print("\n\n--- STANDARD ESTIMATE (for comparison) ---\n")
        std_estimate = calculate_takeoff(measurements)
        print_estimate(std_estimate)

    else:
        # --- MANUAL MODE: user enters all counts ---
        print("\n[DRAINS & SCUPPERS]")
        drains = _input_int("Number of roof drains")
        scuppers = _input_int("Number of overflow scuppers")

        print("\n[EQUIPMENT & PENETRATIONS]")
        mech = _input_int("Mechanical units (RTUs)")
        sleepers = _input_int("Sleeper curbs")
        vents = _input_int("Vent hoods")
        gas = _input_int("Gas penetrations")
        elec = _input_int("Electrical penetrations")
        plumb = _input_int("Plumbing vents")

        print("\n[OPTIONAL OVERRIDES - press Enter to use full roof area]")
        taper_raw = input("  Tapered insulation area (sqft) [full roof]: ").strip()
        taper = float(taper_raw) if taper_raw else None
        ballast_raw = input("  Ballast area (sqft) [full roof]: ").strip()
        ballast = float(ballast_raw) if ballast_raw else None

        measurements = RoofMeasurements(
            total_roof_area_sqft=area,
            perimeter_lf=perim,
            parapet_length_lf=par_len,
            parapet_height_ft=par_ht,
            roof_drain_count=drains,
            scupper_count=scuppers,
            mechanical_unit_count=mech,
            sleeper_curb_count=sleepers,
            vent_hood_count=vents,
            gas_penetration_count=gas,
            electrical_penetration_count=elec,
            plumbing_vent_count=plumb,
            tapered_area_sqft=taper,
            ballast_area_sqft=ballast,
        )

        print(f"\nCalculating estimate...\n")
        estimate = calculate_takeoff(measurements)
        print_estimate(estimate)

    if json_output:
        export_json(estimate, json_output)


if __name__ == "__main__":
    main()
